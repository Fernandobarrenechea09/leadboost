import streamlit as st
from supabase import create_client
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, time as _dtime
from urllib.parse import quote
import pandas as pd
import altair as alt
import calendar as _cal

# ======================================================
# PAGE CONFIG
# ======================================================
st.set_page_config(
    page_title='LEADBOOST DASHBOARD',
    page_icon='#',
    layout='wide',
    initial_sidebar_state='collapsed'
)

# ======================================================
# THEME STATE
# ======================================================
if 'theme' not in st.session_state:
    st.session_state.theme = 'light'

# ======================================================
# THEME COLORS
# ======================================================
if st.session_state.theme == 'light':
    BG = '#F2EDE2'
    PANEL = '#FAF5E9'
    BORDER = '#D8D0BC'
    TEXT = '#1C1C1A'
    TEXT_DIM = '#8A8473'
    ACCENT = '#1C1C1A'
    ACCENT_WARM = '#C4633F'
    CHART_FILL = '#E8C9A8'
    CHART_LINE = '#C4633F'
    HOT_BG = '#EDD5CC'
    WARM_BG = '#E8DCC0'
    COLD_BG = '#D6D9CD'
    WARM_ACCENT = '#CC9A33'
else:
    BG = '#18150F'
    PANEL = '#221E16'
    BORDER = '#3A342A'
    TEXT = '#EDE6D4'
    TEXT_DIM = '#8A8473'
    ACCENT = '#EDE6D4'
    ACCENT_WARM = '#D47A4D'
    CHART_FILL = '#4A3826'
    CHART_LINE = '#D47A4D'
    HOT_BG = '#3A2820'
    WARM_BG = '#332B1E'
    COLD_BG = '#252420'
    WARM_ACCENT = '#DDB04A'

# ======================================================
# CSS
# ======================================================
css = '''
<style>
@import url('https://fonts.googleapis.com/css2?family=Fraunces:ital,opsz,wght@0,9..144,300;0,9..144,400;1,9..144,300;1,9..144,400&family=JetBrains+Mono:wght@400;500&family=Inter:wght@400;500;600&display=swap');

*, *::before, *::after { box-sizing: border-box; }

html, body, [class*="css"], .stApp, .main {
    font-family: Inter, sans-serif;
    background-color: BG_COLOR !important;
    color: TEXT_COLOR !important;
}
.stApp { background-color: BG_COLOR !important; }
.block-container {
    padding-top: 0.55rem !important;
    padding-bottom: 2.5rem !important;
    max-width: 1280px !important;
}

/* Hide Streamlit chrome */
header { visibility: hidden; height: 0 !important; }
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
[data-testid="collapsedControl"] { display: none !important; }
.vega-embed details { display: none !important; }
.vega-embed summary { display: none !important; }
.vega-embed .vega-actions { display: none !important; }

/* Slim scrollbar */
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: BG_COLOR; }
::-webkit-scrollbar-thumb { background: BORDER_COLOR; border-radius: 10px; }
::-webkit-scrollbar-thumb:hover { background: TEXT_DIM_COLOR; }

/* ---- ANIMATIONS ---- */
@keyframes lb-hot-pulse {
    0%   { transform:scale(1);   opacity:.65; }
    70%  { transform:scale(2.4); opacity:0; }
    100% { transform:scale(2.4); opacity:0; }
}
@keyframes lb-fade-up {
    from { opacity:0; transform:translateY(8px); }
    to   { opacity:1; transform:translateY(0); }
}
@keyframes lb-fade-in {
    from { opacity:0; }
    to   { opacity:1; }
}
@keyframes lb-grow-x {
    from { transform: scaleX(0); opacity: 0.4; }
    to   { transform: scaleX(1); opacity: 1; }
}

/* Selection color */
::selection { background: ACCENT_WARM_COLOR; color: BG_COLOR; }

/* ---- TOPBAR ---- */
.lb-topbar-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.58rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    padding-top: 14px;
    padding-bottom: 2px;
}

/* ---- HERO ---- */
.lb-hero {
    border: 1px solid BORDER_COLOR;
    background: PANEL_COLOR;
    border-radius: 16px;
    padding: 34px 38px;
    margin-bottom: 10px;
    display: flex;
    justify-content: space-between;
    align-items: flex-end;
    gap: 28px;
}
.lb-hero-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    letter-spacing: 0.24em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-bottom: 12px;
}
.lb-hero-title {
    font-family: Fraunces, serif;
    font-size: 4rem;
    color: TEXT_COLOR;
    line-height: 0.95;
    font-style: italic;
    font-weight: 300;
    letter-spacing: -0.035em;
}
.lb-hero-accent { color: ACCENT_WARM_COLOR; }

/* ---- PRIORITY CARD ---- */
.priority-card {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-left: 4px solid ACCENT_WARM_COLOR;
    border-radius: 14px;
    padding: 4px 26px;
    margin-top: 6px;
    margin-bottom: 14px;
    animation: lb-fade-up 0.4s ease both;
}
.priority-header {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    padding: 18px 0 14px 0;
    border-bottom: 1px solid BORDER_COLOR;
    margin-bottom: 4px;
}
.priority-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.58rem;
    letter-spacing: 0.22em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
    font-weight: 500;
}
.priority-meta {
    font-family: Fraunces, serif;
    font-size: 1.05rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_DIM_COLOR;
}
.priority-scroll {
    max-height: 360px;
    overflow-y: auto;
    padding-right: 6px;
    margin: 0 -2px;
}
.priority-scroll::-webkit-scrollbar { width: 4px; }
.priority-scroll::-webkit-scrollbar-track { background: transparent; }
.priority-scroll::-webkit-scrollbar-thumb { background: BORDER_COLOR; border-radius: 3px; }
.priority-scroll::-webkit-scrollbar-thumb:hover { background: TEXT_DIM_COLOR; }
.priority-row {
    display: flex;
    align-items: center;
    gap: 22px;
    padding: 16px 0;
    border-bottom: 1px solid BORDER_COLOR;
    transition: padding-left 0.15s, background 0.15s;
}
.priority-row:last-child { border-bottom: none; }
.priority-row:hover { padding-left: 8px; }
.priority-num {
    font-family: Fraunces, serif;
    font-size: 1.7rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_DIM_COLOR;
    letter-spacing: -0.02em;
    line-height: 1;
    min-width: 42px;
    flex-shrink: 0;
}
.priority-row:first-of-type .priority-num {
    color: ACCENT_WARM_COLOR;
}
.priority-body {
    flex: 1;
    min-width: 0;
}
.priority-name {
    font-family: Fraunces, serif;
    font-size: 1.45rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    line-height: 1.15;
    margin-bottom: 6px;
    letter-spacing: -0.015em;
}
.priority-reason {
    font-family: JetBrains Mono, monospace;
    font-size: 0.52rem;
    letter-spacing: 0.18em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.priority-reason .pr-accent {
    color: ACCENT_WARM_COLOR;
    font-weight: 700;
}
.priority-reason .pr-accent-hot { color: ACCENT_WARM_COLOR; font-weight: 700; }
.priority-reason .pr-accent-warm { color: WARM_ACCENT_COLOR; font-weight: 700; }
.priority-reason .pr-accent-cold { color: TEXT_DIM_COLOR; font-weight: 700; }
.priority-action {
    font-family: JetBrains Mono, monospace;
    font-size: 0.54rem;
    letter-spacing: 0.18em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
    text-decoration: none;
    border: 1px solid ACCENT_WARM_COLOR;
    padding: 8px 18px;
    border-radius: 24px;
    transition: background 0.15s, color 0.15s, transform 0.15s;
    flex-shrink: 0;
    white-space: nowrap;
}
.priority-action:hover {
    background: ACCENT_WARM_COLOR;
    color: BG_COLOR;
    transform: translateY(-1px);
}
/* Empty state for priority card */
.priority-empty-card {
    border-left-color: BORDER_COLOR;
}
.priority-empty {
    padding: 36px 4px 32px 4px;
    text-align: left;
}
.priority-empty-title {
    font-family: Fraunces, serif;
    font-size: 2rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    letter-spacing: -0.025em;
    line-height: 1.1;
    margin-bottom: 10px;
}
.priority-empty-sub {
    font-family: JetBrains Mono, monospace;
    font-size: 0.54rem;
    letter-spacing: 0.18em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    max-width: 540px;
}

/* ---- BRIEFING PROSE ---- */
.lb-briefing {
    font-family: Fraunces, serif;
    font-size: 1.15rem;
    font-weight: 300;
    color: TEXT_COLOR;
    margin-top: 18px;
    line-height: 1.55;
    max-width: 680px;
    letter-spacing: -0.005em;
    animation: lb-fade-in 0.5s ease both 0.15s;
}
.b-num {
    font-style: italic;
    color: TEXT_COLOR;
    font-weight: 400;
}
.b-num-accent {
    font-style: italic;
    color: ACCENT_WARM_COLOR;
    font-weight: 400;
}
.b-place {
    font-style: italic;
    color: ACCENT_WARM_COLOR;
    font-weight: 400;
}
.lb-hero-right {
    display: flex;
    flex-direction: column;
    align-items: flex-end;
    gap: 8px;
}
.lb-hero-ring-wrap {
    display: flex;
    align-items: center;
    gap: 12px;
}
.lb-hero-ring-stats {
    display: flex;
    flex-direction: column;
    gap: 4px;
    text-align: right;
}
.lb-hero-ring-stat {
    font-family: JetBrains Mono, monospace;
    font-size: 0.54rem;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: TEXT_DIM_COLOR;
}
.lb-hero-ring-stat.hot { color: ACCENT_WARM_COLOR; }
.lb-hero-ring-stat.warm { color: WARM_ACCENT_COLOR; }
.lb-hero-ring-stat.cold { color: TEXT_DIM_COLOR; }
.lb-hero-date {
    font-family: JetBrains Mono, monospace;
    font-size: 0.58rem;
    color: TEXT_DIM_COLOR;
    letter-spacing: 0.12em;
    text-align: right;
    line-height: 1.7;
}

/* ---- SECTION LABEL ---- */
.section-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.58rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    padding-top: 28px;
    margin-bottom: 14px;
    display: flex;
    justify-content: space-between;
    align-items: baseline;
}
.section-label-left {
    display: flex;
    align-items: baseline;
    gap: 10px;
}
.section-num {
    font-family: Fraunces, serif;
    font-style: italic;
    font-weight: 300;
    font-size: 1.15rem;
    color: TEXT_COLOR;
    letter-spacing: -0.02em;
    line-height: 1;
}
.section-sep {
    color: BORDER_COLOR;
    font-weight: 300;
    font-size: 0.9rem;
    font-family: Fraunces, serif;
    font-style: italic;
}
.section-count {
    font-family: Fraunces, serif;
    font-size: 0.95rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_DIM_COLOR;
}

/* ---- STAT CARDS ---- */
.stat-card {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 20px 22px;
    height: 168px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    transition: border-color 0.2s, transform 0.2s;
}
.stat-card:hover {
    border-color: TEXT_DIM_COLOR;
    transform: translateY(-1px);
}
/* legacy classes kept for backwards-compat — no special border now */
.stat-card-default, .stat-card-hot, .stat-card-warm, .stat-card-cold { }
.stat-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.54rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.stat-value {
    font-family: Fraunces, serif;
    font-size: 3rem;
    color: TEXT_COLOR;
    line-height: 1;
    font-style: italic;
    font-weight: 300;
    letter-spacing: -0.035em;
    font-feature-settings: "tnum";
    font-variant-numeric: tabular-nums;
}
.stat-value-accent {
    font-family: Fraunces, serif;
    font-size: 3rem;
    color: ACCENT_WARM_COLOR;
    line-height: 1;
    font-style: italic;
    font-weight: 300;
    letter-spacing: -0.035em;
    font-feature-settings: "tnum";
    font-variant-numeric: tabular-nums;
}
.stat-spark {
    margin: 4px 0 -2px 0;
    height: 24px;
    display: flex;
    align-items: flex-end;
}
.stat-sub {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    color: TEXT_DIM_COLOR;
    letter-spacing: 0.14em;
    text-transform: uppercase;
}

/* ---- ASYMMETRIC KPI HERO ---- */
.big-stat-card {
    position: relative;
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 14px;
    padding: 28px 32px 22px 32px;
    height: 100%;
    min-height: 280px;
    overflow: hidden;
    transition: border-color 0.2s, transform 0.2s;
    animation: lb-fade-up 0.4s ease both;
}
.big-stat-card:hover { border-color: TEXT_DIM_COLOR; }
.big-stat-bg {
    position: absolute;
    left: 0;
    right: 0;
    bottom: 0;
    height: 60%;
    opacity: 0.5;
    pointer-events: none;
}
.big-stat-content { position: relative; z-index: 2; }
.big-stat-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-bottom: 10px;
}
.big-stat-num {
    font-family: Fraunces, serif;
    font-size: 6.5rem;
    color: TEXT_COLOR;
    line-height: 0.92;
    font-style: italic;
    font-weight: 300;
    letter-spacing: -0.04em;
    font-feature-settings: "tnum";
    font-variant-numeric: tabular-nums;
    margin-bottom: 8px;
}
.big-stat-delta {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    letter-spacing: 0.18em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
}
.big-stat-delta-dim {
    color: TEXT_DIM_COLOR;
}

.mini-stat-stack {
    display: flex;
    flex-direction: column;
    gap: 10px;
    height: 100%;
}
.mini-stat {
    flex: 1;
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 16px 20px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 14px;
    transition: border-color 0.2s, transform 0.2s;
    animation: lb-fade-up 0.4s ease both;
}
.mini-stat:hover { border-color: TEXT_DIM_COLOR; transform: translateY(-1px); }
.mini-stat-left { display: flex; flex-direction: column; gap: 4px; min-width: 0; }
.mini-stat-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.52rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.mini-stat-value {
    font-family: Fraunces, serif;
    font-size: 2.2rem;
    color: TEXT_COLOR;
    line-height: 0.95;
    font-style: italic;
    font-weight: 300;
    letter-spacing: -0.03em;
    font-feature-settings: "tnum";
    font-variant-numeric: tabular-nums;
}
.mini-stat-value-accent { color: ACCENT_WARM_COLOR; }
.mini-stat-value-warm { color: WARM_ACCENT_COLOR; }
.mini-stat-value-cold { color: TEXT_DIM_COLOR; }
.mini-stat-right {
    display: flex;
    flex-direction: column;
    align-items: flex-end;
    gap: 4px;
    min-width: 90px;
    flex-shrink: 0;
}
.mini-stat-pct {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.mini-stat-spark {
    width: 100%;
    max-width: 110px;
}

/* ---- CALENDAR HEATMAP ---- */
.heatmap-wrap {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 24px 28px;
    animation: lb-fade-up 0.4s ease both;
}
.heatmap-header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    margin-bottom: 22px;
    padding-bottom: 16px;
    border-bottom: 1px solid BORDER_COLOR;
    gap: 24px;
}
.heatmap-title-block { flex: 1; min-width: 0; }
.heatmap-meta-block {
    display: flex;
    flex-direction: column;
    align-items: flex-end;
    gap: 4px;
    flex-shrink: 0;
}
.heatmap-title {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-bottom: 8px;
}
.heatmap-subtitle {
    font-family: Fraunces, serif;
    font-size: 0.95rem;
    color: TEXT_COLOR;
    font-style: italic;
    font-weight: 300;
    line-height: 1.4;
    max-width: 480px;
    letter-spacing: -0.005em;
}
.heatmap-meta {
    font-family: Fraunces, serif;
    font-size: 1.1rem;
    color: TEXT_COLOR;
    font-style: italic;
    font-weight: 300;
}
.heatmap-busiest {
    font-family: JetBrains Mono, monospace;
    font-size: 0.46rem;
    letter-spacing: 0.18em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
}
.heatmap-svg-wrap {
    width: 100%;
    max-width: 100%;
    overflow: hidden;
    padding: 6px 0;
}
.heatmap-legend {
    display: flex;
    align-items: center;
    justify-content: flex-end;
    gap: 8px;
    margin-top: 14px;
    font-family: JetBrains Mono, monospace;
    font-size: 0.46rem;
    letter-spacing: 0.16em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.heatmap-legend-cells { display: inline-flex; gap: 3px; }
.heatmap-legend-cell {
    width: 11px;
    height: 11px;
    border-radius: 2px;
    background: ACCENT_WARM_COLOR;
}

/* ---- CALENDAR (month view) ---- */
.cal-wrap {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 22px 24px;
    margin-top: 8px;
    animation: lb-fade-up 0.4s ease both;
}
.cal-header {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    margin-bottom: 16px;
    padding-bottom: 12px;
    border-bottom: 1px solid BORDER_COLOR;
}
.cal-month-title {
    font-family: Fraunces, serif;
    font-size: 1.6rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    letter-spacing: -0.02em;
    line-height: 1;
}
.cal-month-meta {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.18em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.cal-grid {
    display: grid;
    grid-template-columns: repeat(7, 1fr);
    gap: 6px;
}
.cal-dow {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    padding: 6px 0 10px 0;
    text-align: center;
}
.cal-cell {
    background: BG_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 6px;
    padding: 8px 10px;
    min-height: 78px;
    display: flex;
    flex-direction: column;
    transition: border-color 0.15s, transform 0.15s, background 0.15s;
    position: relative;
}
.cal-cell:hover {
    border-color: TEXT_DIM_COLOR;
    transform: translateY(-1px);
}
.cal-cell.other-month { opacity: 0.32; }
.cal-cell.today {
    border-color: ACCENT_WARM_COLOR;
    border-width: 2px;
    padding: 7px 9px;
}
.cal-cell.today::after {
    content: 'TODAY';
    position: absolute;
    top: 6px;
    right: 8px;
    font-family: JetBrains Mono, monospace;
    font-size: 0.42rem;
    letter-spacing: 0.2em;
    color: ACCENT_WARM_COLOR;
}
.cal-cell.focus {
    border-color: ACCENT_COLOR;
    border-width: 2px;
    padding: 7px 9px;
    box-shadow: 0 0 0 3px BG_COLOR, 0 0 0 4px ACCENT_COLOR;
}
.cal-cell.focus.today {
    border-color: ACCENT_WARM_COLOR;
    box-shadow: 0 0 0 3px BG_COLOR, 0 0 0 4px ACCENT_WARM_COLOR;
}
.cal-cell.focus::before {
    content: '▾';
    position: absolute;
    top: 4px;
    left: 8px;
    font-size: 0.6rem;
    color: ACCENT_COLOR;
    line-height: 1;
}
.cal-cell.has-hot {
    background: linear-gradient(180deg, PANEL_COLOR 0%, BG_COLOR 60%);
}
.cal-day-num {
    font-family: Fraunces, serif;
    font-size: 1.05rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    line-height: 1;
    margin-bottom: 6px;
    letter-spacing: -0.01em;
    font-feature-settings: "tnum";
}
.cal-cell.today .cal-day-num { color: ACCENT_WARM_COLOR; }
.cal-dots {
    display: flex;
    gap: 3px;
    margin-bottom: 4px;
    flex-wrap: wrap;
}
.cal-dot {
    width: 6px;
    height: 6px;
    border-radius: 50%;
}
.cal-dot.dot-hot { background: ACCENT_WARM_COLOR; }
.cal-dot.dot-warm { background: WARM_ACCENT_COLOR; }
.cal-dot.dot-cold {
    background: TEXT_DIM_COLOR;
    opacity: 0.55;
}
.cal-count {
    font-family: JetBrains Mono, monospace;
    font-size: 0.48rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-top: auto;
}
.cal-empty {
    margin-top: auto;
    font-family: JetBrains Mono, monospace;
    font-size: 0.42rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    opacity: 0.4;
    text-transform: uppercase;
}

/* Calendar nav buttons — slimmer than default */
.cal-nav-row { margin-bottom: 4px; }
.cal-nav-row div.stButton > button {
    padding: 6px 16px !important;
    font-size: 0.5rem !important;
    letter-spacing: 0.14em !important;
    border-radius: 6px !important;
    min-height: 32px !important;
    height: 32px !important;
    width: 100% !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    line-height: 1.2 !important;
}

/* ---- DAY DETAIL CARD (focus day expansion) ---- */
.day-detail-card {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-left: 3px solid ACCENT_COLOR;
    border-radius: 12px;
    padding: 24px 28px;
    margin-top: 12px;
    animation: lb-fade-up 0.3s ease both;
}
.day-detail-card.is-today { border-left-color: ACCENT_WARM_COLOR; }
.day-detail-header {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    margin-bottom: 18px;
    padding-bottom: 12px;
    border-bottom: 1px solid BORDER_COLOR;
}
.day-detail-title {
    font-family: Fraunces, serif;
    font-size: 1.6rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    letter-spacing: -0.025em;
    line-height: 1;
}
.day-detail-badge {
    font-family: JetBrains Mono, monospace;
    font-size: 0.48rem;
    letter-spacing: 0.22em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
    padding: 4px 10px;
    border: 1px solid ACCENT_WARM_COLOR;
    border-radius: 16px;
}
.day-detail-section { margin-bottom: 20px; }
.day-detail-section:last-child { margin-bottom: 0; }
.day-detail-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.52rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-bottom: 10px;
}
.day-detail-event {
    display: flex;
    gap: 18px;
    padding: 12px 0;
    border-bottom: 1px solid BORDER_COLOR;
}
.day-detail-event:last-child { border-bottom: none; }
.day-detail-event-time {
    font-family: JetBrains Mono, monospace;
    font-size: 0.62rem;
    letter-spacing: 0.16em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
    min-width: 76px;
    padding-top: 3px;
    flex-shrink: 0;
}
.day-detail-event-body { flex: 1; }
.day-detail-event-title {
    font-family: Fraunces, serif;
    font-size: 1.1rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    margin-bottom: 5px;
    letter-spacing: -0.005em;
}
.day-detail-event-loc {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.day-detail-event-notes {
    font-family: Inter, sans-serif;
    font-size: 0.82rem;
    color: TEXT_COLOR;
    opacity: 0.88;
    line-height: 1.5;
    margin-top: 4px;
}
.day-detail-lead {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 9px 0;
    border-bottom: 1px solid BORDER_COLOR;
}
.day-detail-lead:last-child { border-bottom: none; }
.day-detail-lead-name {
    font-family: Fraunces, serif;
    font-size: 0.98rem;
    font-style: italic;
    color: TEXT_COLOR;
    flex: 1;
    font-weight: 300;
}
.day-detail-lead-meta {
    font-family: JetBrains Mono, monospace;
    font-size: 0.46rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.day-detail-empty {
    font-family: Fraunces, serif;
    font-size: 1.1rem;
    font-style: italic;
    color: TEXT_DIM_COLOR;
    padding: 16px 0 8px 0;
    font-weight: 300;
}

/* View-day picker — match other labels */
.view-day-wrap {
    margin-top: 12px;
    margin-bottom: 4px;
}
/* Faux-label spacer so the TODAY button aligns vertically with the date input below its label */
.vd-faux-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.7rem;
    color: transparent;
    line-height: 1.4;
    height: 1.4em;
    margin-bottom: 4px;
    user-select: none;
}
/* Style the TODAY button to feel paired with the date input */
.vd-today-btn-wrap div.stButton > button,
.vd-today-btn-wrap [data-testid="stButton"] button {
    width: 100% !important;
    background: PANEL_COLOR !important;
    color: TEXT_COLOR !important;
    border: 1px solid BORDER_COLOR !important;
    border-radius: 8px !important;
    padding: 0 22px !important;
    height: 38px !important;
    min-height: 38px !important;
    font-family: JetBrains Mono, monospace !important;
    font-size: 0.58rem !important;
    letter-spacing: 0.16em !important;
    text-transform: uppercase !important;
    box-shadow: none !important;
    transition: border-color 0.2s, color 0.2s, background 0.2s !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    line-height: 1.2 !important;
}
.vd-today-btn-wrap div.stButton > button:hover,
.vd-today-btn-wrap [data-testid="stButton"] button:hover {
    border-color: ACCENT_WARM_COLOR !important;
    color: ACCENT_WARM_COLOR !important;
    background: PANEL_COLOR !important;
}

/* Event chips inside calendar cells */
.cal-event-chip {
    display: block;
    font-family: JetBrains Mono, monospace;
    font-size: 0.46rem;
    letter-spacing: 0.08em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
    margin-top: 4px;
    padding: 2px 5px;
    background: ACCENT_WARM_COLOR;
    background: linear-gradient(90deg, rgba(196,99,63,0.18), rgba(196,99,63,0.06));
    border-left: 2px solid ACCENT_WARM_COLOR;
    border-radius: 2px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    line-height: 1.3;
}
.cal-event-extra {
    font-family: JetBrains Mono, monospace;
    font-size: 0.42rem;
    color: TEXT_DIM_COLOR;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    margin-top: 2px;
}

/* ---- AGENDA ---- */
.agenda-wrap {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 24px 28px;
    animation: lb-fade-up 0.4s ease both;
}
.agenda-group { margin-bottom: 24px; }
.agenda-group:last-child { margin-bottom: 0; }
.agenda-period-row {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    padding-bottom: 10px;
    margin-bottom: 12px;
    border-bottom: 1px solid BORDER_COLOR;
}
.agenda-period {
    font-family: JetBrains Mono, monospace;
    font-size: 0.6rem;
    letter-spacing: 0.24em;
    color: TEXT_COLOR;
    text-transform: uppercase;
}
.agenda-period-count {
    font-family: Fraunces, serif;
    font-size: 1.05rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_DIM_COLOR;
}
.agenda-empty {
    font-family: Fraunces, serif;
    font-size: 0.9rem;
    font-style: italic;
    color: TEXT_DIM_COLOR;
    padding: 4px 0;
}
.agenda-item {
    display: flex;
    gap: 16px;
    padding: 12px 4px;
    border-bottom: 1px solid BORDER_COLOR;
    transition: padding-left 0.15s;
}
.agenda-item:last-child { border-bottom: none; }
.agenda-item:hover { padding-left: 10px; }
.agenda-item-time {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.16em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
    min-width: 64px;
    padding-top: 4px;
    flex-shrink: 0;
}
.agenda-item-body { flex: 1; min-width: 0; }
.agenda-item-title {
    font-family: Fraunces, serif;
    font-size: 1.05rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    line-height: 1.3;
    margin-bottom: 4px;
    letter-spacing: -0.005em;
}
.agenda-item-meta {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.agenda-item-notes {
    font-family: Inter, sans-serif;
    font-size: 0.78rem;
    color: TEXT_COLOR;
    opacity: 0.85;
    line-height: 1.4;
    margin-top: 4px;
}
.agenda-item-actions {
    flex-shrink: 0;
    display: flex;
    align-items: center;
    gap: 8px;
}

/* Lead activity inside agenda — different style */
.agenda-lead-row {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 10px 4px;
    border-bottom: 1px solid BORDER_COLOR;
    transition: padding-left 0.15s;
}
.agenda-lead-row:hover { padding-left: 10px; }
.agenda-lead-name {
    font-family: Fraunces, serif;
    font-size: 0.95rem;
    font-style: italic;
    color: TEXT_COLOR;
    font-weight: 300;
    flex: 1;
}
.agenda-lead-tag {
    font-family: JetBrains Mono, monospace;
    font-size: 0.46rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}

/* Add Event form panel */
.add-event-panel {
    background: BG_COLOR;
    border: 1px dashed BORDER_COLOR;
    border-radius: 10px;
    padding: 18px 22px;
    margin: 8px 0 16px 0;
}
.add-event-title {
    font-family: JetBrains Mono, monospace;
    font-size: 0.58rem;
    letter-spacing: 0.22em;
    color: ACCENT_WARM_COLOR;
    text-transform: uppercase;
    margin-bottom: 12px;
}

/* ---- PULL-QUOTE CALLOUTS ---- */
.pull-quote {
    display: flex;
    align-items: stretch;
    gap: 24px;
    padding: 32px 8px 32px 4px;
    margin: 14px 0 6px 0;
    animation: lb-fade-up 0.5s ease both;
}
.pull-quote-rule {
    width: 3px;
    background: ACCENT_WARM_COLOR;
    border-radius: 2px;
    flex-shrink: 0;
}
.pull-quote-content { flex: 1; }
.pull-quote-text {
    font-family: Fraunces, serif;
    font-style: italic;
    font-weight: 300;
    font-size: 2.2rem;
    color: TEXT_COLOR;
    line-height: 1.2;
    letter-spacing: -0.025em;
    margin-bottom: 10px;
    max-width: 880px;
}
.pull-quote-text-accent { color: ACCENT_WARM_COLOR; font-weight: 400; }
.pull-quote-attr {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}

/* ---- STAGGERED ANIMATION ---- */
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(1) .stat-card,
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(1) .mini-stat,
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(1) .resp-card { animation-delay: 0ms; }
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(2) .stat-card,
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(2) .mini-stat,
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(2) .resp-card { animation-delay: 70ms; }
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(3) .stat-card,
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(3) .mini-stat,
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(3) .resp-card { animation-delay: 140ms; }
[data-testid="stHorizontalBlock"] [data-testid="stColumn"]:nth-child(4) .stat-card { animation-delay: 210ms; }

/* ---- FUNNEL STRIP ---- */
.funnel-strip {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 14px 20px;
    display: flex;
    align-items: center;
    margin-bottom: 8px;
}
.funnel-step { flex: 1; text-align: center; }
.funnel-label {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.2em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    margin-bottom: 3px;
}
.funnel-num {
    font-family: Fraunces, serif;
    font-size: 1.7rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    line-height: 1;
    font-feature-settings: "tnum";
}
.funnel-pct {
    font-family: JetBrains Mono, monospace;
    font-size: 0.48rem;
    color: TEXT_DIM_COLOR;
    letter-spacing: 0.1em;
    margin-top: 3px;
}
.funnel-arrow {
    font-size: 0.9rem;
    color: BORDER_COLOR;
    padding: 0 8px;
    flex-shrink: 0;
    padding-bottom: 14px;
}

/* ---- CHART PANELS ---- */
.chart-panel {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 14px 18px 6px 18px;
    margin-bottom: 8px;
}
.chart-header {
    display: flex;
    justify-content: space-between;
    align-items: baseline;
    margin-bottom: 8px;
    padding-bottom: 8px;
    border-bottom: 1px solid BORDER_COLOR;
}
.chart-title {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    letter-spacing: 0.22em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.chart-big-num {
    font-family: Fraunces, serif;
    font-size: 1rem;
    color: TEXT_COLOR;
    font-style: italic;
    font-weight: 300;
}

/* ---- ZONE LEADERBOARD ---- */
.zone-board {
    display: flex;
    flex-direction: column;
    gap: 9px;
    padding: 4px 0 6px 0;
}
.zone-row {
    display: flex;
    align-items: center;
    gap: 10px;
}
.zone-name {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    letter-spacing: 0.08em;
    color: TEXT_COLOR;
    min-width: 110px;
    text-transform: uppercase;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.zone-bar-wrap {
    flex: 1;
    height: 4px;
    background: BORDER_COLOR;
    border-radius: 10px;
    overflow: hidden;
}
.zone-bar {
    height: 100%;
    background: CHART_LINE_COLOR;
    border-radius: 10px;
}
.zone-count {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    color: TEXT_DIM_COLOR;
    min-width: 18px;
    text-align: right;
}

/* ---- RESPONSE TIME CARD ---- */
.resp-card {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 15px 18px;
    height: 124px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    border-top: 3px solid BORDER_COLOR;
}

/* ---- BADGES ---- */
.badge {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    letter-spacing: 0.18em;
    padding: 3px 8px;
    border-radius: 20px;
    border: 1px solid BORDER_COLOR;
    text-transform: uppercase;
    font-weight: 500;
    color: TEXT_COLOR;
    display: inline-block;
}
.score-hot {
    background: HOT_BG_COLOR;
    border-color: ACCENT_WARM_COLOR;
    color: ACCENT_WARM_COLOR;
}
.score-warm {
    background: WARM_BG_COLOR;
    border-color: WARM_ACCENT_COLOR;
    color: WARM_ACCENT_COLOR;
}
.score-cold {
    background: COLD_BG_COLOR;
    border-color: BORDER_COLOR;
    color: TEXT_DIM_COLOR;
}

/* ---- LEAD CARDS ---- */
.lead-card {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
    padding: 22px 26px;
    margin-bottom: 8px;
    animation: lb-fade-up 0.3s ease both;
    transition: border-color 0.2s, transform 0.2s;
}
.lead-card:hover {
    border-color: TEXT_DIM_COLOR;
    transform: translateY(-1px);
}
.lead-card-hot { border-color: ACCENT_WARM_COLOR55; }
.lead-card-hot:hover { border-color: ACCENT_WARM_COLOR; }
.lead-card-warm { border-color: WARM_ACCENT_COLOR55; }
.lead-card-warm:hover { border-color: WARM_ACCENT_COLOR; }
.lead-card-cold { border-color: BORDER_COLOR; }
.lead-top {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 14px;
    padding-bottom: 12px;
    border-bottom: 1px solid BORDER_COLOR;
}
.lead-name {
    font-family: Fraunces, serif;
    font-size: 1.55rem;
    font-style: italic;
    color: TEXT_COLOR;
    font-weight: 300;
    letter-spacing: -0.02em;
}
.lead-badges { display: flex; gap: 5px; align-items: center; }
.lead-grid {
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 8px 22px;
    font-family: JetBrains Mono, monospace;
    font-size: 0.7rem;
    color: TEXT_COLOR;
    margin-bottom: 10px;
}
.field-label {
    color: TEXT_DIM_COLOR;
    font-size: 0.5rem;
    letter-spacing: 0.2em;
    text-transform: uppercase;
    display: block;
    margin-bottom: 2px;
}
.lead-footer {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-top: 8px;
    padding-top: 9px;
    border-top: 1px solid BORDER_COLOR;
}
.lead-ts {
    font-family: JetBrains Mono, monospace;
    font-size: 0.5rem;
    color: TEXT_DIM_COLOR;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}
.wa-btn {
    display: inline-block;
    background: ACCENT_COLOR;
    color: BG_COLOR;
    padding: 6px 14px;
    border-radius: 8px;
    text-decoration: none;
    font-family: JetBrains Mono, monospace;
    font-weight: 500;
    font-size: 0.56rem;
    letter-spacing: 0.18em;
    text-transform: uppercase;
    transition: opacity 0.15s;
}
.wa-btn:hover { opacity: 0.8; }
.resp-time {
    font-family: JetBrains Mono, monospace;
    font-size: 0.54rem;
    color: TEXT_DIM_COLOR;
    letter-spacing: 0.1em;
}

/* ---- INPUTS ---- */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stDateInput input,
.stTimeInput input,
[data-testid="stDateInput"] input,
[data-testid="stTimeInput"] input {
    background: PANEL_COLOR !important;
    color: TEXT_COLOR !important;
    border: 1px solid BORDER_COLOR !important;
    border-radius: 8px !important;
    font-family: Inter, sans-serif !important;
    font-size: 0.86rem !important;
    transition: border-color 0.2s, box-shadow 0.2s, background 0.2s !important;
}

/* Form input labels — darker, mono caps, editorial */
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] label,
.stTextInput label,
.stTextArea label,
.stDateInput label,
.stTimeInput label,
.stSelectbox label,
.stNumberInput label {
    font-family: JetBrains Mono, monospace !important;
    font-size: 0.58rem !important;
    letter-spacing: 0.22em !important;
    color: TEXT_COLOR !important;
    text-transform: uppercase !important;
    font-weight: 500 !important;
    opacity: 0.85;
}
.stTextInput > div > div > input:hover,
.stTextArea > div > div > textarea:hover,
.stDateInput input:hover,
.stTimeInput input:hover {
    border-color: TEXT_DIM_COLOR !important;
}
/* DateInput / TimeInput popovers + outer container */
.stDateInput > div,
.stTimeInput > div {
    background: transparent !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: ACCENT_WARM_COLOR !important;
    box-shadow: 0 0 0 2px ACCENT_WARM_COLOR33 !important;
    outline: none !important;
}
.stSelectbox > div > div {
    background: PANEL_COLOR !important;
    border: 1px solid BORDER_COLOR !important;
    border-radius: 8px !important;
    color: TEXT_COLOR !important;
}
.stTextInput label, .stSelectbox label, .stTextArea label {
    color: TEXT_DIM_COLOR !important;
    font-family: JetBrains Mono, monospace !important;
    font-size: 0.52rem !important;
    letter-spacing: 0.22em !important;
    text-transform: uppercase !important;
}

/* ---- BUTTONS ---- */
/* Ghost / secondary buttons (default for all st.button calls) */
div.stButton > button {
    background: transparent;
    color: TEXT_DIM_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 8px;
    padding: 8px 24px;
    font-family: JetBrains Mono, monospace;
    font-weight: 500;
    font-size: 0.58rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    transition: all 0.15s;
    white-space: nowrap;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    line-height: 1.2;
    min-height: 38px;
}
div.stButton > button p,
div.stButton > button span,
div.stButton > button div {
    margin: 0 !important;
    padding: 0 !important;
    font-family: inherit !important;
    font-size: inherit !important;
    letter-spacing: inherit !important;
    line-height: inherit !important;
}
div.stButton > button:hover:not(:disabled) {
    border-color: TEXT_COLOR;
    color: TEXT_COLOR;
}
/* Primary solid button (use type='primary' in Python) */
div.stButton > button[kind="primary"] {
    background: ACCENT_COLOR !important;
    color: BG_COLOR !important;
    border: none !important;
    opacity: 1;
}
div.stButton > button[kind="primary"]:hover {
    opacity: 0.85 !important;
}
/* Active status button (disabled=True) — dark accent, restrained */
div.stButton > button:disabled {
    opacity: 1 !important;
    background: ACCENT_COLOR !important;
    color: BG_COLOR !important;
    border: none !important;
    cursor: default !important;
}
/* Download (export) button — ghost */
div.stDownloadButton > button {
    background: transparent;
    color: TEXT_DIM_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 8px;
    padding: 8px 24px;
    font-family: JetBrains Mono, monospace;
    font-weight: 500;
    font-size: 0.58rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    white-space: nowrap;
    transition: all 0.15s;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    line-height: 1.2;
    min-height: 38px;
}
div.stDownloadButton > button p,
div.stDownloadButton > button span,
div.stDownloadButton > button div {
    margin: 0 !important;
    padding: 0 !important;
    font-family: inherit !important;
    font-size: inherit !important;
    letter-spacing: inherit !important;
}
div.stDownloadButton > button:hover {
    border-color: TEXT_COLOR;
    color: TEXT_COLOR;
}
/* Theme toggle pill — first stHorizontalBlock, last column */
[data-testid="stHorizontalBlock"]:first-child [data-testid="stColumn"]:last-child div.stButton > button {
    border-radius: 20px !important;
    font-size: 0.5rem !important;
    padding: 3px 10px !important;
    height: auto !important;
    min-height: unset !important;
    line-height: 1.4 !important;
    width: auto !important;
    min-width: unset !important;
    letter-spacing: 0.14em !important;
    box-shadow: none !important;
}

/* ---- EMPTY STATE ---- */
.empty-state-rich {
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-left: 4px solid BORDER_COLOR;
    border-radius: 14px;
    padding: 56px 36px 52px 36px;
    margin-top: 8px;
    animation: lb-fade-up 0.4s ease both;
}
.empty-state-title {
    font-family: Fraunces, serif;
    font-size: 2.4rem;
    font-style: italic;
    font-weight: 300;
    color: TEXT_COLOR;
    letter-spacing: -0.03em;
    line-height: 1.05;
    margin-bottom: 14px;
}
.empty-state-sub {
    font-family: JetBrains Mono, monospace;
    font-size: 0.56rem;
    letter-spacing: 0.18em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
    max-width: 560px;
    line-height: 1.6;
}

.empty-state {
    text-align: center;
    padding: 40px 20px;
    font-family: Fraunces, serif;
    font-style: italic;
    font-size: 1.2rem;
    font-weight: 300;
    color: TEXT_DIM_COLOR;
    background: PANEL_COLOR;
    border: 1px solid BORDER_COLOR;
    border-radius: 12px;
}

/* ---- NOTES AREA ---- */
.stTextArea { margin-top: 4px !important; }
.stTextArea > div > div > textarea { font-size: 0.8rem !important; }

/* Reduce Streamlit column gap */
[data-testid="stHorizontalBlock"] { gap: 0.5rem !important; }

/* ---- COLLAPSIBLE SECTION CARDS (st.expander) ---- */
[data-testid="stExpander"],
div[data-testid="stExpander"] {
    background: PANEL_COLOR !important;
    border: 1px solid BORDER_COLOR !important;
    border-radius: 12px !important;
    margin-bottom: 10px !important;
    transition: border-color 0.2s !important;
    overflow: hidden !important;
}
[data-testid="stExpander"]:hover,
div[data-testid="stExpander"]:hover {
    border-color: TEXT_DIM_COLOR !important;
}
[data-testid="stExpander"] details {
    background: transparent !important;
    border: none !important;
}
[data-testid="stExpander"] details > summary,
[data-testid="stExpander"] summary {
    font-family: JetBrains Mono, monospace !important;
    font-size: 0.62rem !important;
    letter-spacing: 0.22em !important;
    color: TEXT_COLOR !important;
    text-transform: uppercase !important;
    padding: 20px 24px !important;
    background: transparent !important;
    border: none !important;
    border-radius: 12px !important;
    cursor: pointer !important;
    list-style: none !important;
    transition: color 0.2s, background 0.15s !important;
}
[data-testid="stExpander"] details > summary:hover,
[data-testid="stExpander"] summary:hover {
    color: ACCENT_WARM_COLOR !important;
    background: rgba(196, 99, 63, 0.04) !important;
}
[data-testid="stExpander"] details > summary p,
[data-testid="stExpander"] summary p {
    font-family: JetBrains Mono, monospace !important;
    font-size: 0.62rem !important;
    letter-spacing: 0.22em !important;
    text-transform: uppercase !important;
    margin: 0 !important;
}
[data-testid="stExpander"] details[open] > summary {
    border-bottom: 1px solid BORDER_COLOR !important;
    border-bottom-left-radius: 0 !important;
    border-bottom-right-radius: 0 !important;
}
[data-testid="stExpander"] svg {
    fill: TEXT_DIM_COLOR !important;
    color: TEXT_DIM_COLOR !important;
    transition: transform 0.2s !important;
}
[data-testid="stExpander"] [data-testid="stExpanderDetails"],
[data-testid="stExpander"] details[open] > div {
    padding: 18px 24px 22px 24px !important;
    background: transparent !important;
    animation: lb-fade-in 0.25s ease both !important;
}
/* Reset inner section labels — they're now redundant inside expanders */
[data-testid="stExpander"] .section-label {
    padding-top: 0 !important;
    margin-bottom: 12px !important;
}

/* ---- LEAD EXPANDERS — per-lead score-colored left border ---- */
.lead-expander-wrap { margin-bottom: 8px; }
.lead-expander-wrap [data-testid="stExpander"],
.lead-expander-wrap div[data-testid="stExpander"] {
    border-left-width: 3px !important;
    border-radius: 10px !important;
    margin-bottom: 0 !important;
}
.lead-expander-wrap.lead-hot [data-testid="stExpander"] {
    border-left-color: ACCENT_WARM_COLOR !important;
}
.lead-expander-wrap.lead-warm [data-testid="stExpander"] {
    border-left-color: WARM_ACCENT_COLOR !important;
}
.lead-expander-wrap.lead-cold [data-testid="stExpander"] {
    border-left-color: BORDER_COLOR !important;
}
.lead-expander-wrap [data-testid="stExpander"] summary {
    padding: 16px 22px !important;
    font-family: Inter, sans-serif !important;
    font-size: 0.78rem !important;
    letter-spacing: 0.04em !important;
    text-transform: none !important;
    color: TEXT_COLOR !important;
}
.lead-expander-wrap [data-testid="stExpander"] summary p {
    font-family: Inter, sans-serif !important;
    font-size: 0.78rem !important;
    letter-spacing: 0.04em !important;
    text-transform: none !important;
}
.lead-expander-wrap [data-testid="stExpander"] summary strong,
.lead-expander-wrap [data-testid="stExpander"] summary p strong {
    font-family: Fraunces, serif !important;
    font-style: italic !important;
    font-weight: 300 !important;
    font-size: 1.25rem !important;
    letter-spacing: -0.01em !important;
    color: TEXT_COLOR !important;
}

/* ---- FOLLOW-UP SCROLLABLE LIST ---- */
.follow-up-list {
    max-height: 312px;
    overflow-y: auto;
    padding-right: 6px;
    margin: 0 -2px;
}
.follow-up-list::-webkit-scrollbar { width: 3px; }
.follow-up-list::-webkit-scrollbar-track { background: transparent; }
.follow-up-list::-webkit-scrollbar-thumb { background: BORDER_COLOR; border-radius: 2px; }
.follow-up-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 12px 4px;
    border-bottom: 1px solid BORDER_COLOR;
    transition: background 0.15s, padding-left 0.15s;
    cursor: default;
}
.follow-up-row:last-child { border-bottom: none; }
.follow-up-row:hover { padding-left: 10px; }
.follow-up-name {
    font-family: Fraunces, serif;
    font-size: 1.05rem;
    font-style: italic;
    color: TEXT_COLOR;
    font-weight: 300;
    margin-bottom: 3px;
    letter-spacing: -0.01em;
}
.follow-up-meta {
    font-family: JetBrains Mono, monospace;
    font-size: 0.48rem;
    letter-spacing: 0.14em;
    color: TEXT_DIM_COLOR;
    text-transform: uppercase;
}
.follow-up-right {
    display: flex;
    flex-direction: column;
    align-items: flex-end;
    gap: 4px;
}
.follow-up-ago {
    font-family: JetBrains Mono, monospace;
    font-size: 0.46rem;
    color: TEXT_DIM_COLOR;
    letter-spacing: 0.12em;
    text-transform: uppercase;
}
</style>
'''

# Replace longer/specific placeholders first to avoid substring collisions.
# BG_COLOR is inside HOT_BG_COLOR / WARM_BG_COLOR / COLD_BG_COLOR.
# ACCENT_COLOR is inside ACCENT_WARM_COLOR and WARM_ACCENT_COLOR.
# TEXT_COLOR is inside TEXT_DIM_COLOR.
css = css.replace('HOT_BG_COLOR', HOT_BG)
css = css.replace('WARM_BG_COLOR', WARM_BG)
css = css.replace('COLD_BG_COLOR', COLD_BG)
css = css.replace('WARM_ACCENT_COLOR', WARM_ACCENT)
css = css.replace('ACCENT_WARM_COLOR', ACCENT_WARM)
css = css.replace('TEXT_DIM_COLOR', TEXT_DIM)
css = css.replace('CHART_LINE_COLOR', CHART_LINE)
css = css.replace('BG_COLOR', BG)
css = css.replace('PANEL_COLOR', PANEL)
css = css.replace('BORDER_COLOR', BORDER)
css = css.replace('TEXT_COLOR', TEXT)
css = css.replace('ACCENT_COLOR', ACCENT)
st.markdown(css, unsafe_allow_html=True)

# ======================================================
# TOP BAR
# ======================================================
now_time = datetime.now().strftime('%H:%M')
today_short = datetime.now().strftime('%Y%m%d').upper()

top_l, top_m, top_r = st.columns([3, 5, 1])
with top_l:
    st.markdown('<div class="lb-topbar-label">// LEADBOOST OS &nbsp; V3.0</div>', unsafe_allow_html=True)
with top_m:
    st.markdown('<div class="lb-topbar-label" style="text-align:center;">BOLIVIA &middot; ' + today_short + ' &middot; ' + now_time + '</div>', unsafe_allow_html=True)
with top_r:
    if st.button('DARK' if st.session_state.theme == 'light' else 'LIGHT', key='theme_toggle_dash'):
        st.session_state.theme = 'dark' if st.session_state.theme == 'light' else 'light'
        st.rerun()

# ======================================================
# PASSWORD
# ======================================================
OWNER_PASSWORD = 'leadboost2024'
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    today_full = datetime.now().strftime('%A, %B %d, %Y').upper()
    hero = '<div class="lb-hero">'
    hero += '<div><div class="lb-hero-label">// DASHBOARD &middot; AGENTS ONLY</div>'
    hero += '<div class="lb-hero-title">Private <span class="lb-hero-accent">access</span>.</div></div>'
    hero += '<div class="lb-hero-right"><div class="lb-hero-date">BOLIVIA<br>' + now_time + '</div></div>'
    hero += '</div>'
    st.markdown(hero, unsafe_allow_html=True)
    pw = st.text_input('PASSWORD', type='password')
    if st.button('ENTER  //', type='primary'):
        if pw == OWNER_PASSWORD:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error('Incorrect password.')
    st.stop()

# ======================================================
# DB FUNCTIONS
# ======================================================
@st.cache_resource
def get_supabase():
    return create_client(st.secrets['SUPABASE_URL'], st.secrets['SUPABASE_KEY'])

def load_leads():
    try:
        return get_supabase().table('leads').select('*').order('id', desc=True).execute().data
    except Exception as e:
        st.error('Error: ' + str(e))
        return []

def update_status(lead_id, new_status):
    try:
        data = {'status': new_status}
        if new_status == 'Contactado':
            data['contacted_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        get_supabase().table('leads').update(data).eq('id', lead_id).execute()
    except Exception as e:
        st.error('Error: ' + str(e))

# ======================================================
# EVENTS (agenda items)
# ======================================================
def fetch_events():
    try:
        res = get_supabase().table('events').select('*').order('event_date').order('event_time').execute()
        return res.data or []
    except Exception:
        # Table may not exist yet — fail silently
        return []

def add_event(event_date, event_time, title, location, notes, lead_id):
    try:
        data = {
            'event_date': event_date,
            'title': title,
        }
        if event_time:
            data['event_time'] = event_time
        if location:
            data['location'] = location
        if notes:
            data['notes'] = notes
        if lead_id:
            data['lead_id'] = lead_id
        get_supabase().table('events').insert(data).execute()
        return True
    except Exception as e:
        st.error('Error adding event: ' + str(e))
        return False

def delete_event(event_id):
    try:
        get_supabase().table('events').delete().eq('id', event_id).execute()
        return True
    except Exception as e:
        st.error('Error deleting event: ' + str(e))
        return False

def update_event(event_id, event_date, event_time, title, location, notes, lead_id):
    try:
        data = {
            'event_date': event_date,
            'title': title,
            'event_time': event_time,  # may be None — that's OK to clear time
            'location': location,
            'notes': notes,
            'lead_id': lead_id,
        }
        get_supabase().table('events').update(data).eq('id', event_id).execute()
        return True
    except Exception as e:
        st.error('Error updating event: ' + str(e))
        return False

def get_minutes(lead):
    try:
        t1 = datetime.strptime(lead.get('timestamp', ''), '%Y-%m-%d %H:%M')
        t2 = datetime.strptime(lead.get('contacted_at', ''), '%Y-%m-%d %H:%M')
        m = int((t2 - t1).total_seconds() / 60)
        return m if m >= 0 else None
    except:
        return None

def fmt(minutes):
    if minutes is None: return '-'
    if minutes < 60: return str(minutes) + ' min'
    if minutes < 1440: return str(round(minutes / 60, 1)) + 'h'
    return str(round(minutes / 1440, 1)) + 'd'

def generate_excel(leads):
    wb = Workbook()
    ws = wb.active
    ws.title = 'LeadBoost'
    h_fill = PatternFill('solid', fgColor='1C1C1A')
    hot_f = PatternFill('solid', fgColor='EDD5CC')
    warm_f = PatternFill('solid', fgColor='E8DCC0')
    cold_f = PatternFill('solid', fgColor='D6D9CD')
    alt_f = PatternFill('solid', fgColor='FAF5E9')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    c = Alignment(horizontal='center', vertical='center')
    l = Alignment(horizontal='left', vertical='center')
    headers = ['#', 'Fecha', 'Nombre', 'Telefono', 'Tipo', 'Zona', 'Presupuesto', 'Plazo', 'Score', 'Estado', 'T. Respuesta']
    widths = [5, 18, 22, 15, 18, 18, 16, 12, 14, 14, 16]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = h_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = c
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    for rn, lead in enumerate(leads, 2):
        score = lead.get('score', 'COLD')
        if score == 'HOT':
            row_fill = hot_f
        elif score == 'WARM':
            row_fill = warm_f
        else:
            row_fill = cold_f if rn % 2 == 0 else alt_f
        minutes = get_minutes(lead)
        row = [rn - 1, lead.get('timestamp', ''), lead.get('name', ''), lead.get('phone', ''),
               lead.get('property_type', ''), lead.get('area', ''), lead.get('budget', ''),
               lead.get('timeline', ''), score, lead.get('status', 'Nuevo'), fmt(minutes)]
        for cn, val in enumerate(row, 1):
            cell = ws.cell(row=rn, column=cn, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = c if cn in [1, 7, 8, 9, 10, 11] else l
            cell.font = Font(bold=True, size=10) if cn == 3 else Font(size=10)
        ws.row_dimensions[rn].height = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ======================================================
# COMPONENT HELPERS
# ======================================================
_sec_n = [0]
def sec_label(name, right_html=''):
    _sec_n[0] += 1
    num = str(_sec_n[0]).zfill(2)
    return ('<div class="section-label">'
            '<span class="section-label-left">'
            '<span class="section-num">' + num + '</span>'
            '<span class="section-sep">/</span>'
            '<span>' + name + '</span>'
            '</span>'
            + right_html + '</div>')

def compute_priority(leads_list, max_items=20):
    """Score each lead by urgency and return the top max_items needing action."""
    now = datetime.now()
    scored = []
    for l in leads_list:
        status = l.get('status', 'Nuevo')
        # Skip closed/visited — those don't need next action
        if status in ('Visitado', 'Cerrado'):
            continue
        score = l.get('score', 'COLD')
        ts = l.get('timestamp', '') or ''
        try:
            ts_dt = datetime.strptime(ts[:16], '%Y-%m-%d %H:%M')
            age_days = max(0, (now - ts_dt).days)
        except:
            age_days = 0

        # Score-based weight (HOT > WARM > COLD)
        score_weight = {'HOT': 100, 'WARM': 55, 'COLD': 25}.get(score, 20)
        # Status modifier
        if status == 'Nuevo':
            urgency = score_weight + age_days * 3
            if age_days == 0:
                reason_html = '<span class="pr-accent-' + score.lower() + '">' + score + '</span> &middot; NEW LEAD &middot; ACT TODAY'
            elif age_days == 1:
                reason_html = '<span class="pr-accent-' + score.lower() + '">' + score + '</span> &middot; NEW &middot; 1 DAY WAITING'
            else:
                reason_html = '<span class="pr-accent-' + score.lower() + '">' + score + '</span> &middot; NEW &middot; ' + str(age_days) + ' DAYS WAITING'
        else:  # Contactado
            # Days since contacted (or fall back to received age)
            contacted_at = (l.get('contacted_at') or '').strip()
            contact_days = age_days
            if contacted_at:
                try:
                    c_dt = datetime.strptime(contacted_at[:16], '%Y-%m-%d %H:%M')
                    contact_days = max(0, (now - c_dt).days)
                except:
                    pass
            urgency = (score_weight - 25) + contact_days * 2
            if contact_days == 0:
                reason_html = '<span class="pr-accent-' + score.lower() + '">' + score + '</span> &middot; CONTACTED TODAY &middot; KEEP MOMENTUM'
            elif contact_days == 1:
                reason_html = '<span class="pr-accent-' + score.lower() + '">' + score + '</span> &middot; FOLLOW-UP DUE &middot; 1 DAY SINCE CONTACT'
            else:
                reason_html = '<span class="pr-accent-' + score.lower() + '">' + score + '</span> &middot; FOLLOW-UP DUE &middot; ' + str(contact_days) + ' DAYS SINCE CONTACT'
        # Score rank: HOT first, then WARM, then COLD (lower = higher priority)
        score_rank = {'HOT': 0, 'WARM': 1, 'COLD': 2}.get(score, 3)
        scored.append({'lead': l, 'urgency': urgency, 'reason_html': reason_html, 'score_rank': score_rank})

    # Primary sort: score (HOT > WARM > COLD). Secondary: urgency (older/more urgent first within score).
    scored.sort(key=lambda x: (x['score_rank'], -x['urgency']))
    return scored[:max_items]

def expander_label(name, teaser=''):
    """Build an editorial expander summary string. Increments the section counter."""
    _sec_n[0] += 1
    num = str(_sec_n[0]).zfill(2)
    base = '//  ' + num + '   ·   ' + name
    if teaser:
        base += '   ·   ' + teaser
    return base

def daily_counts(leads_list, score_filter=None, days=30):
    today = datetime.now().date()
    dates = [(today - timedelta(days=days-1-i)).strftime('%Y-%m-%d') for i in range(days)]
    counts = {d: 0 for d in dates}
    for l in leads_list:
        if score_filter and l.get('score') != score_filter:
            continue
        ts = (l.get('timestamp') or '')[:10]
        if ts in counts:
            counts[ts] += 1
    return [counts[d] for d in dates]

def sparkline_svg(values, color, width=160, height=22):
    if not values:
        return ''
    max_v = max(values) if max(values) > 0 else 1
    n = len(values)
    pts = []
    for i, v in enumerate(values):
        x = (i / (n - 1) * width) if n > 1 else width / 2
        y = height - (v / max_v * (height - 4)) - 2
        pts.append((round(x, 1), round(y, 1)))
    line = 'M' + ' L'.join(str(p[0]) + ',' + str(p[1]) for p in pts)
    area = line + ' L' + str(pts[-1][0]) + ',' + str(height) + ' L' + str(pts[0][0]) + ',' + str(height) + ' Z'
    return ('<svg width="100%" height="' + str(height) + '" viewBox="0 0 ' + str(width) + ' ' + str(height) + '" preserveAspectRatio="none" style="display:block;">'
            '<path d="' + area + '" fill="' + color + '" opacity="0.12"/>'
            '<path d="' + line + '" fill="none" stroke="' + color + '" stroke-width="1.4" stroke-linecap="round" stroke-linejoin="round" opacity="0.85"/>'
            '</svg>')

def time_ago(ts_str):
    try:
        t = datetime.strptime(ts_str, '%Y-%m-%d %H:%M')
        delta = datetime.now() - t
        mins = int(delta.total_seconds() / 60)
        if mins < 60: return str(mins) + 'M AGO'
        if mins < 1440: return str(mins // 60) + 'H AGO'
        return str(mins // 1440) + 'D AGO'
    except:
        return ''

def heatmap_svg(leads_list, color, weeks=12):
    today = datetime.now().date()
    # Snap end-date to the most recent Sunday so each column is a clean Mon-Sun week
    today_dow = today.weekday()  # 0=Mon, 6=Sun
    days_to_sunday = (6 - today_dow) % 7
    end = today + timedelta(days=days_to_sunday)
    days = weeks * 7
    start = end - timedelta(days=days-1)  # this is a Monday

    counts = {}
    for l in leads_list:
        ts = (l.get('timestamp') or '')[:10]
        try:
            d = datetime.strptime(ts, '%Y-%m-%d').date()
            counts[d] = counts.get(d, 0) + 1
        except:
            continue
    max_count = max(counts.values()) if counts else 0

    cell = 18
    gap = 5
    label_col_w = 36   # space for day-of-week labels on the left
    label_row_h = 22   # space for month labels along the top

    grid_w = weeks * (cell + gap)
    grid_h = 7 * (cell + gap)
    total_w = label_col_w + grid_w
    total_h = label_row_h + grid_h

    svg = '<svg width="100%" height="' + str(total_h) + '" viewBox="0 0 ' + str(total_w) + ' ' + str(total_h) + '" preserveAspectRatio="xMidYMid meet" style="display:block;">'

    # Month labels along the top — only when month changes between weeks
    last_month = None
    for wi in range(weeks):
        sample = start + timedelta(days=wi * 7)
        if sample.month != last_month:
            x = label_col_w + wi * (cell + gap)
            svg += ('<text x="' + str(x) + '" y="' + str(label_row_h - 8)
                    + '" font-family="JetBrains Mono, monospace" font-size="9"'
                    + ' fill="' + TEXT_DIM + '" letter-spacing="1.6">'
                    + sample.strftime('%b').upper() + '</text>')
            last_month = sample.month

    # Day-of-week labels (left) — show Mon / Wed / Fri / Sun
    day_labels = ['MON', '', 'WED', '', 'FRI', '', 'SUN']
    for di, lbl in enumerate(day_labels):
        if not lbl:
            continue
        y = label_row_h + di * (cell + gap) + cell - 5
        svg += ('<text x="0" y="' + str(y)
                + '" font-family="JetBrains Mono, monospace" font-size="8.5"'
                + ' fill="' + TEXT_DIM + '" letter-spacing="1.2">' + lbl + '</text>')

    # Cells
    for wi in range(weeks):
        for di in range(7):
            i = wi * 7 + di
            day = start + timedelta(days=i)
            count = counts.get(day, 0)
            is_future = day > today
            if is_future:
                # Future cells: very faint, dashed feel
                opacity = 0.04
            elif max_count > 0 and count > 0:
                opacity = 0.25 + (count / max_count) * 0.75
            else:
                opacity = 0.10
            x = label_col_w + wi * (cell + gap)
            y = label_row_h + di * (cell + gap)
            svg += ('<rect x="' + str(x) + '" y="' + str(y)
                    + '" width="' + str(cell) + '" height="' + str(cell)
                    + '" rx="3" fill="' + color + '" opacity="' + str(round(opacity, 2)) + '">'
                    + '<title>' + day.strftime('%a %b %d') + ': ' + str(count) + ' lead' + ('s' if count != 1 else '') + '</title></rect>')

    svg += '</svg>'
    return svg

def pull_quote(text_html, attribution=''):
    html = '<div class="pull-quote">'
    html += '<div class="pull-quote-rule"></div>'
    html += '<div class="pull-quote-content">'
    html += '<div class="pull-quote-text">' + text_html + '</div>'
    if attribution:
        html += '<div class="pull-quote-attr">' + attribution + '</div>'
    html += '</div></div>'
    return html

def mini_stat(label, value, pct_text, sparkline_html='', accent=False, score=None):
    base = 'mini-stat-value'
    if score == 'HOT' or accent:
        val_class = base + ' mini-stat-value-accent'
    elif score == 'WARM':
        val_class = base + ' mini-stat-value-warm'
    elif score == 'COLD':
        val_class = base + ' mini-stat-value-cold'
    else:
        val_class = base
    html = '<div class="mini-stat">'
    html += '<div class="mini-stat-left">'
    html += '<div class="mini-stat-label">' + label + '</div>'
    html += '<div class="' + val_class + '">' + str(value) + '</div>'
    html += '</div>'
    html += '<div class="mini-stat-right">'
    if sparkline_html:
        html += '<div class="mini-stat-spark">' + sparkline_html + '</div>'
    html += '<div class="mini-stat-pct">' + pct_text + '</div>'
    html += '</div>'
    html += '</div>'
    return html

def big_stat_card(label, value, delta_text, sparkline_full=''):
    html = '<div class="big-stat-card">'
    if sparkline_full:
        html += '<div class="big-stat-bg">' + sparkline_full + '</div>'
    html += '<div class="big-stat-content">'
    html += '<div class="big-stat-label">' + label + '</div>'
    html += '<div class="big-stat-num">' + str(value) + '</div>'
    html += '<div class="big-stat-delta">' + delta_text + '</div>'
    html += '</div>'
    html += '</div>'
    return html

def stat_card(label, value, sub, accent=False, top_class='stat-card-default', sparkline=''):
    val_class = 'stat-value-accent' if accent else 'stat-value'
    html = '<div class="stat-card ' + top_class + '">'
    html += '<div class="stat-label">' + label + '</div>'
    html += '<div class="' + val_class + '">' + str(value) + '</div>'
    if sparkline:
        html += '<div class="stat-spark">' + sparkline + '</div>'
    html += '<div class="stat-sub">' + sub + '</div>'
    html += '</div>'
    return html

def resp_card(label, value, sub, accent=False):
    val_class = 'stat-value-accent' if accent else 'stat-value'
    html = '<div class="resp-card">'
    html += '<div class="stat-label">' + label + '</div>'
    html += '<div class="' + val_class + '">' + str(value) + '</div>'
    html += '<div class="stat-sub">' + sub + '</div>'
    html += '</div>'
    return html

def score_ring_svg(hot, warm, cold, total):
    r = 22
    pi = 3.14159265
    c = round(2 * pi * r, 2)

    def seg(length, color, start):
        gap = round(c - length, 2)
        # Correct dashoffset: c - start places the segment at the right position
        offset = round(c - start, 2)
        return ('<circle cx="28" cy="28" r="' + str(r) + '" fill="none" stroke="' + color + '" stroke-width="5"'
                ' stroke-dasharray="' + str(length) + ' ' + str(gap) + '"'
                ' stroke-dashoffset="' + str(offset) + '"/>')

    svg = '<svg width="56" height="56" style="transform:rotate(-90deg);flex-shrink:0;">'
    svg += '<circle cx="28" cy="28" r="' + str(r) + '" fill="none" stroke="' + BORDER + '" stroke-width="5"/>'

    if total > 0:
        hot_len = round((hot / total) * c, 2)
        warm_len = round((warm / total) * c, 2)
        cold_len = round(c - hot_len - warm_len, 2)
        if cold_len > 0:
            svg += seg(cold_len, TEXT_DIM, hot_len + warm_len)
        if warm_len > 0:
            svg += seg(warm_len, WARM_ACCENT, hot_len)
        if hot_len > 0:
            svg += seg(hot_len, ACCENT_WARM, 0)

    svg += '</svg>'
    return svg

def funnel_strip(leads):
    statuses = ['Nuevo', 'Contactado', 'Visitado', 'Cerrado']
    counts = {s: sum(1 for l in leads if l.get('status', 'Nuevo') == s) for s in statuses}
    total = len(leads)
    html = '<div class="funnel-strip">'
    for i, s in enumerate(statuses):
        n = counts[s]
        pct = str(round(n / total * 100)) + '%' if total > 0 else '0%'
        html += '<div class="funnel-step">'
        html += '<div class="funnel-label">' + s + '</div>'
        html += '<div class="funnel-num">' + str(n) + '</div>'
        html += '<div class="funnel-pct">' + pct + ' of total</div>'
        html += '</div>'
        if i < len(statuses) - 1:
            html += '<div class="funnel-arrow">&#x2192;</div>'
    html += '</div>'
    return html

def zones_leaderboard(leads, top_n=6):
    from collections import Counter
    zone_counts = Counter(
        l.get('area', '').strip() for l in leads if l.get('area', '').strip()
    )
    top = zone_counts.most_common(top_n)
    if not top:
        return ''
    max_count = top[0][1]
    html = '<div class="zone-board">'
    for zone, count in top:
        bar_pct = round(count / max_count * 100)
        html += '<div class="zone-row">'
        html += '<div class="zone-name">' + zone + '</div>'
        html += '<div class="zone-bar-wrap"><div class="zone-bar" style="width:' + str(bar_pct) + '%"></div></div>'
        html += '<div class="zone-count">' + str(count) + '</div>'
        html += '</div>'
    html += '</div>'
    return html

# ======================================================
# LOAD DATA
# ======================================================
leads = load_leads()
total = len(leads)
hot = sum(1 for l in leads if l.get('score') == 'HOT')
warm = sum(1 for l in leads if l.get('score') == 'WARM')
cold = sum(1 for l in leads if l.get('score') == 'COLD')

def pct(n):
    if total == 0: return '0%'
    return str(round(n / total * 100)) + '%'

# ======================================================
# HERO
# ======================================================
hour = datetime.now().hour
if hour < 12:
    greeting_word = 'morning'
elif hour < 18:
    greeting_word = 'afternoon'
else:
    greeting_word = 'evening'

today_full = datetime.now().strftime('%A, %B %d, %Y').upper()
weekday_name = datetime.now().strftime('%A')
ring = score_ring_svg(hot, warm, cold, total)

# Briefing data
pending_count = sum(1 for l in leads if l.get('status', 'Nuevo') in ('Nuevo', 'Contactado'))
resp_minutes_list = [get_minutes(l) for l in leads]
resp_minutes_list = [m for m in resp_minutes_list if m is not None]
avg_resp = int(sum(resp_minutes_list) / len(resp_minutes_list)) if resp_minutes_list else None

from collections import Counter as _C
_zone_counts = _C(l.get('area', '').strip() for l in leads if l.get('area', '').strip())
_top_zone_pair = _zone_counts.most_common(1)
top_zone_name = _top_zone_pair[0][0] if _top_zone_pair else None

# Build briefing prose
briefing_parts = []
if hot > 0:
    briefing_parts.append('You have <span class="b-num-accent">' + str(hot) + '</span> hot lead' + ('s' if hot != 1 else ''))
elif total > 0:
    briefing_parts.append('You have <span class="b-num">' + str(total) + '</span> lead' + ('s' if total != 1 else '') + ' on file')
if pending_count > 0:
    briefing_parts.append('<span class="b-num">' + str(pending_count) + '</span> follow-up' + ('s' if pending_count != 1 else '') + ' pending')
if avg_resp is not None:
    briefing_parts.append('an average response of <span class="b-num">' + fmt(avg_resp) + '</span>')

briefing_main = ''
if briefing_parts:
    if len(briefing_parts) == 1:
        briefing_main = briefing_parts[0] + '.'
    elif len(briefing_parts) == 2:
        briefing_main = briefing_parts[0] + ' and ' + briefing_parts[1] + '.'
    else:
        briefing_main = ', '.join(briefing_parts[:-1]) + ', and ' + briefing_parts[-1] + '.'
else:
    briefing_main = 'No leads yet today — your dashboard is ready.'

briefing_zone = ''
if top_zone_name:
    briefing_zone = ' Top zone: <span class="b-place">' + top_zone_name + '</span>.'

hero = '<div class="lb-hero">'
hero += '<div style="flex:1;min-width:0;">'
hero += '<div class="lb-hero-label">// OPERATIONS PANEL &middot; LIVE</div>'
hero += '<div class="lb-hero-title">Good <span class="lb-hero-accent">' + greeting_word + '</span>.</div>'
hero += '<div class="lb-briefing">' + briefing_main + briefing_zone + '</div>'
hero += '</div>'
hero += '<div class="lb-hero-right">'
hero += '<div class="lb-hero-ring-wrap">'
hero += ring
hero += '<div class="lb-hero-ring-stats">'
hero += '<div class="lb-hero-ring-stat hot">&#x25CF; ' + str(hot) + ' hot</div>'
hero += '<div class="lb-hero-ring-stat warm">&#x25CF; ' + str(warm) + ' warm</div>'
hero += '<div class="lb-hero-ring-stat cold">' + str(cold) + ' cold</div>'
hero += '</div>'
hero += '</div>'
hero += '<div class="lb-hero-date">' + today_full + '<br>LOCAL TIME &middot; ' + now_time + '</div>'
hero += '</div>'
hero += '</div>'
st.markdown(hero, unsafe_allow_html=True)

# ======================================================
# PRIORITY — most urgent next actions for the agent
# ======================================================
if leads:
    _priorities = compute_priority(leads, max_items=20)
    if not _priorities:
        # Empty state — all active leads have been actioned
        _p_empty = '<div class="priority-card priority-empty-card">'
        _p_empty += '<div class="priority-header">'
        _p_empty += '<div class="priority-label">// PRIORITY &middot; WHAT TO DO NOW</div>'
        _p_empty += '<div class="priority-meta">all clear</div>'
        _p_empty += '</div>'
        _p_empty += '<div class="priority-empty">'
        _p_empty += '<div class="priority-empty-title">All caught up.</div>'
        _p_empty += '<div class="priority-empty-sub">No urgent actions right now. Every active lead has been contacted.</div>'
        _p_empty += '</div>'
        _p_empty += '</div>'
        st.markdown(_p_empty, unsafe_allow_html=True)
    if _priorities:
        _p_html = '<div class="priority-card">'
        _p_html += '<div class="priority-header">'
        _p_html += '<div class="priority-label">// PRIORITY &middot; WHAT TO DO NOW</div>'
        _p_html += '<div class="priority-meta">' + str(len(_priorities)) + ' action' + ('s' if len(_priorities) != 1 else '') + '</div>'
        _p_html += '</div>'
        _p_html += '<div class="priority-scroll">'
        for _idx, _p in enumerate(_priorities, 1):
            _pl = _p['lead']
            _p_name = _pl.get('name', '-')
            _p_phone = (_pl.get('phone', '') or '').replace('+', '').replace(' ', '').replace('-', '')
            _p_ptype = _pl.get('property_type', '') or ''
            _p_area = _pl.get('area', '') or ''
            _p_budget = _pl.get('budget', '-')
            _p_wa_msg = 'Hola ' + _p_name + ', soy de la agencia inmobiliaria LeadBoost. Te contactamos porque mostraste interes en ' + _p_ptype + ' en ' + _p_area + '. Tienes un momento para hablar?'
            _p_wa_link = 'https://wa.me/591' + _p_phone + '?text=' + quote(_p_wa_msg)

            _p_html += '<div class="priority-row">'
            _p_html += '<div class="priority-num">' + str(_idx).zfill(2) + '</div>'
            _p_html += '<div class="priority-body">'
            _p_html += '<div class="priority-name">' + _p_name + '</div>'
            _p_reason_meta = _p['reason_html']
            if _p_budget and str(_p_budget) != '-':
                _p_reason_meta += ' &middot; $' + str(_p_budget)
            if _p_area:
                _p_reason_meta += ' &middot; ' + _p_area.upper()
            _p_html += '<div class="priority-reason">' + _p_reason_meta + '</div>'
            _p_html += '</div>'
            _p_html += '<a class="priority-action" href="' + _p_wa_link + '" target="_blank">// WHATSAPP</a>'
            _p_html += '</div>'
        _p_html += '</div>'  # close priority-scroll
        _p_html += '</div>'  # close priority-card
        st.markdown(_p_html, unsafe_allow_html=True)

# ======================================================
# TOP ACTIONS
# ======================================================
btn_refresh, btn_export, _ = st.columns([2, 2, 6])
with btn_refresh:
    if st.button('// REFRESH', key='refresh_btn'):
        st.cache_resource.clear()
        st.rerun()
with btn_export:
    if leads:
        st.download_button(
            label='// EXPORT XLSX',
            data=generate_excel(leads),
            file_name='leadboost_leads.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            key='export_btn'
        )

# ======================================================
# STATS
# ======================================================
st.markdown(sec_label('OVERVIEW', '<span class="section-count">' + str(total) + ' leads total</span>'), unsafe_allow_html=True)

# Sparklines for each score category
big_spark = sparkline_svg(daily_counts(leads), CHART_LINE, width=600, height=72)
hot_spark = sparkline_svg(daily_counts(leads, 'HOT'), ACCENT_WARM, width=110, height=22)
warm_spark = sparkline_svg(daily_counts(leads, 'WARM'), WARM_ACCENT, width=110, height=22)
cold_spark = sparkline_svg(daily_counts(leads, 'COLD'), TEXT_DIM, width=110, height=22)

# Compute "this week" delta for big stat
_today_d = datetime.now().date()
_week_start = _today_d - timedelta(days=6)
_this_week = sum(1 for l in leads if l.get('timestamp', '')[:10] >= _week_start.strftime('%Y-%m-%d'))
delta_text = '<span class="big-stat-delta-dim">LAST 7 DAYS &middot;</span> +' + str(_this_week) + ' NEW'

big_col, mini_col = st.columns([1, 1])
with big_col:
    st.markdown(big_stat_card('// TOTAL LEADS &middot; ALL TIME', total, delta_text, sparkline_full=big_spark), unsafe_allow_html=True)
with mini_col:
    stack = '<div class="mini-stat-stack">'
    stack += mini_stat('// HOT', hot, pct(hot) + ' OF TOTAL', sparkline_html=hot_spark, score='HOT')
    stack += mini_stat('// WARM', warm, pct(warm) + ' OF TOTAL', sparkline_html=warm_spark, score='WARM')
    stack += mini_stat('// COLD', cold, pct(cold) + ' OF TOTAL', sparkline_html=cold_spark, score='COLD')
    stack += '</div>'
    st.markdown(stack, unsafe_allow_html=True)

# ======================================================
# PRE-COMPUTE METRICS FOR EXPANDER TEASERS + PULL-QUOTE
# ======================================================
if leads:
    _today_date = datetime.now().strftime('%Y-%m-%d')
    today_leads = sum(1 for l in leads if l.get('timestamp', '').startswith(_today_date))
    pending_contact = sum(1 for l in leads if l.get('status', 'Nuevo') == 'Nuevo')
    conv_rate = str(round((hot + warm) / total * 100)) + '%' if total > 0 else '0%'
    # Pipeline stage counts
    _stage_keys = ['Nuevo', 'Contactado', 'Visitado', 'Cerrado']
    _stage_counts = {k: sum(1 for l in leads if l.get('status', 'Nuevo') == k) for k in _stage_keys}
else:
    today_leads = 0
    pending_contact = 0
    conv_rate = '0%'
    _stage_counts = {}

# ======================================================
# PULL-QUOTE — always visible, surfaces the most striking metric
# ======================================================
quote_text = None
quote_attr = None
if total > 0:
    conv_n = round((hot + warm) / total * 100)
    if conv_n >= 50:
        quote_text = '<span class="pull-quote-text-accent">' + str(conv_n) + '%</span> of your leads are warm or hot — your best signal yet.'
        quote_attr = '// CONVERSION INSIGHT'
    elif hot >= 3:
        quote_text = '<span class="pull-quote-text-accent">' + str(hot) + ' hot leads</span> in the pipeline. Move fast — response time matters.'
        quote_attr = '// PRIORITY SIGNAL'
    elif top_zone_name and _top_zone_pair and _top_zone_pair[0][1] >= 2:
        quote_text = '<span class="pull-quote-text-accent">' + top_zone_name + '</span> leads the zones with ' + str(_top_zone_pair[0][1]) + ' leads. Worth doubling down.'
        quote_attr = '// MARKET FOCUS'
    elif resp_minutes_list and avg_resp is not None and avg_resp <= 30:
        quote_text = 'Average response time: <span class="pull-quote-text-accent">' + fmt(avg_resp) + '</span>. Fast enough to win.'
        quote_attr = '// PERFORMANCE'
    else:
        quote_text = 'Steady pipeline. <span class="pull-quote-text-accent">' + str(total) + '</span> leads on file, ' + str(pending_count) + ' still waiting on you.'
        quote_attr = '// OPERATIONS'
if quote_text:
    st.markdown(pull_quote(quote_text, quote_attr), unsafe_allow_html=True)

# ======================================================
# CONVERSION FUNNEL
# ======================================================
if leads:
    st.markdown(sec_label('PIPELINE'), unsafe_allow_html=True)
    st.markdown(funnel_strip(leads), unsafe_allow_html=True)

# ======================================================
# CHARTS
# ======================================================
if leads:
    st.markdown(sec_label('ANALYTICS'), unsafe_allow_html=True)
    ch1, ch2 = st.columns(2)

    with ch1:
        dates = [l.get('timestamp', '')[:10] for l in leads if l.get('timestamp')]
        if dates:
            df = pd.DataFrame({'date': dates})
            df['date'] = pd.to_datetime(df['date'])
            daily = df.groupby('date').size().reset_index(name='Leads').sort_values('date')

            base = alt.Chart(daily)

            area_layer = base.mark_area(
                line={'color': CHART_LINE, 'strokeWidth': 2.2},
                color=alt.Gradient(
                    gradient='linear',
                    stops=[
                        alt.GradientStop(color=CHART_FILL, offset=0),
                        alt.GradientStop(color=BG, offset=1)
                    ],
                    x1=1, x2=1, y1=1, y2=0
                ),
                interpolate='monotone',
                opacity=0.95
            ).encode(
                x=alt.X('date:T', axis=alt.Axis(
                    labelFont='JetBrains Mono',
                    labelFontSize=10,
                    labelColor=TEXT_DIM,
                    domainColor=BORDER,
                    tickColor=BORDER,
                    grid=False,
                    title=None,
                    format='%b %d',
                    labelAngle=0,
                    tickCount=6
                )),
                y=alt.Y('Leads:Q', axis=alt.Axis(
                    labelFont='JetBrains Mono',
                    labelFontSize=10,
                    labelColor=TEXT_DIM,
                    domainColor=BORDER,
                    tickColor=BORDER,
                    gridColor=BORDER,
                    gridOpacity=0.4,
                    title=None,
                    tickCount=4
                ))
            )
            dot_layer = base.mark_point(
                filled=True, size=50, color=CHART_LINE, strokeWidth=2, stroke=PANEL
            ).encode(x='date:T', y='Leads:Q')
            label_layer = base.mark_text(
                font='JetBrains Mono', fontSize=10, color=CHART_LINE, dy=-13
            ).encode(x='date:T', y='Leads:Q', text='Leads:Q')

            area_chart = alt.layer(area_layer, dot_layer, label_layer).properties(
                height=180, background='transparent'
            ).configure_view(strokeWidth=0)

            header = '<div class="chart-panel"><div class="chart-header">'
            header += '<div class="chart-title">// LEADS OVER TIME</div>'
            header += '<div class="chart-big-num">' + str(total) + ' total</div>'
            header += '</div>'
            st.markdown(header, unsafe_allow_html=True)
            st.altair_chart(area_chart, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

    with ch2:
        def bucket(b):
            try:
                v = int(str(b).replace(',', '').replace('.', '').replace('$', '').strip())
                if v < 50000: return '< 50k'
                if v < 100000: return '50-100k'
                if v < 150000: return '100-150k'
                return '> 150k'
            except:
                return 'Unknown'

        order = ['< 50k', '50-100k', '100-150k', '> 150k']
        buckets = [bucket(l.get('budget', '0')) for l in leads]
        counts_df = pd.DataFrame({'Rango': buckets})
        counts_df = counts_df[counts_df['Rango'] != 'Unknown']
        if not counts_df.empty:
            counts_data = counts_df['Rango'].value_counts().reindex(order, fill_value=0).reset_index()
            counts_data.columns = ['Rango', 'Leads']

            bar_base = alt.Chart(counts_data)
            bar_layer = bar_base.mark_bar(
                color=CHART_LINE,
                opacity=0.92,
                cornerRadiusTopLeft=3,
                cornerRadiusTopRight=3
            ).encode(
                x=alt.X('Rango:N', sort=order, axis=alt.Axis(
                    labelFont='JetBrains Mono',
                    labelFontSize=10,
                    labelColor=TEXT_DIM,
                    domainColor=BORDER,
                    tickColor=BORDER,
                    title=None,
                    labelAngle=0
                )),
                y=alt.Y('Leads:Q', axis=alt.Axis(
                    labelFont='JetBrains Mono',
                    labelFontSize=10,
                    labelColor=TEXT_DIM,
                    domainColor=BORDER,
                    tickColor=BORDER,
                    gridColor=BORDER,
                    gridOpacity=0.4,
                    title=None,
                    tickCount=4
                ))
            )
            bar_label_layer = bar_base.mark_text(
                font='JetBrains Mono', fontSize=11, color=CHART_LINE, dy=-8
            ).encode(
                x=alt.X('Rango:N', sort=order),
                y='Leads:Q',
                text=alt.Text('Leads:Q')
            ).transform_filter(alt.datum.Leads > 0)

            bar_chart = alt.layer(bar_layer, bar_label_layer).properties(
                height=180, background='transparent'
            ).configure_view(strokeWidth=0)

            header = '<div class="chart-panel"><div class="chart-header">'
            header += '<div class="chart-title">// BUDGET DISTRIBUTION</div>'
            header += '<div class="chart-big-num">USD ranges</div>'
            header += '</div>'
            st.markdown(header, unsafe_allow_html=True)
            st.altair_chart(bar_chart, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

# ======================================================
# CALENDAR — month view with lead activity per day
# ======================================================
if leads:
    _today_d = datetime.now().date()

    if 'cal_year' not in st.session_state:
        st.session_state.cal_year = _today_d.year
    if 'cal_month' not in st.session_state:
        st.session_state.cal_month = _today_d.month

    _cyr = st.session_state.cal_year
    _cmo = st.session_state.cal_month
    _cur_month_label = datetime(_cyr, _cmo, 1).strftime('%B %Y')

    # Build lead activity map (date -> list of leads)
    _cell_leads = {}
    for _l in leads:
        _ts = (_l.get('timestamp') or '')[:10]
        try:
            _d = datetime.strptime(_ts, '%Y-%m-%d').date()
            _cell_leads.setdefault(_d, []).append(_l)
        except:
            continue

    # Fetch events and build event map (date -> list of events)
    events_all = fetch_events()
    _cell_events = {}
    for _e in events_all:
        try:
            _ed = datetime.strptime(str(_e.get('event_date')), '%Y-%m-%d').date()
            _cell_events.setdefault(_ed, []).append(_e)
        except:
            continue

    # Month summary stats
    _month_total = sum(1 for _l in leads if (_l.get('timestamp') or '')[:7] == datetime(_cyr, _cmo, 1).strftime('%Y-%m'))
    _month_hot = sum(1 for _l in leads if (_l.get('timestamp') or '')[:7] == datetime(_cyr, _cmo, 1).strftime('%Y-%m') and _l.get('score') == 'HOT')
    _month_events = sum(1 for _e in events_all if str(_e.get('event_date', ''))[:7] == datetime(_cyr, _cmo, 1).strftime('%Y-%m'))

    # Section label
    st.markdown(sec_label('CALENDAR', '<span class="section-count">' + str(_month_total) + ' leads &middot; ' + str(_month_events) + ' events this month</span>'), unsafe_allow_html=True)

    # Month title + nav
    title_html = '<div class="cal-header">'
    title_html += '<div class="cal-month-title">' + _cur_month_label + '</div>'
    title_html += '<div class="cal-month-meta">// ' + str(_month_total) + ' LEADS &middot; ' + str(_month_hot) + ' HOT &middot; ' + str(_month_events) + ' EVENTS</div>'
    title_html += '</div>'

    # Start the calendar wrapper
    st.markdown('<div class="cal-wrap">' + title_html, unsafe_allow_html=True)

    # Navigation row
    st.markdown('<div class="cal-nav-row">', unsafe_allow_html=True)
    nav_prev, nav_today, nav_next, _nav_sp = st.columns([1, 1, 1, 6])
    with nav_prev:
        if st.button('< PREV', key='cal_prev_btn'):
            if _cmo == 1:
                st.session_state.cal_month = 12
                st.session_state.cal_year = _cyr - 1
            else:
                st.session_state.cal_month = _cmo - 1
            st.rerun()
    with nav_today:
        if st.button('TODAY', key='cal_today_btn'):
            st.session_state.cal_year = _today_d.year
            st.session_state.cal_month = _today_d.month
            st.rerun()
    with nav_next:
        if st.button('NEXT >', key='cal_next_btn'):
            if _cmo == 12:
                st.session_state.cal_month = 1
                st.session_state.cal_year = _cyr + 1
            else:
                st.session_state.cal_month = _cmo + 1
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    # Focus day state
    if 'focus_day_val' not in st.session_state:
        st.session_state.focus_day_val = _today_d
    _focus_d = st.session_state.focus_day_val

    # Build the grid (Mon-first)
    _c_obj = _cal.Calendar(firstweekday=0)
    _grid_dates = list(_c_obj.itermonthdates(_cyr, _cmo))

    grid_html = '<div class="cal-grid">'
    # Day-of-week headers
    for _dn in ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']:
        grid_html += '<div class="cal-dow">' + _dn + '</div>'
    # Day cells
    for _d in _grid_dates:
        classes = ['cal-cell']
        if _d.month != _cmo:
            classes.append('other-month')
        if _d == _today_d:
            classes.append('today')
        if _d == _focus_d and _d.month == _cmo:
            classes.append('focus')
        day_leads = _cell_leads.get(_d, [])
        n = len(day_leads)
        hot_n = sum(1 for x in day_leads if x.get('score') == 'HOT')
        warm_n = sum(1 for x in day_leads if x.get('score') == 'WARM')
        cold_n = sum(1 for x in day_leads if x.get('score') == 'COLD')
        if hot_n > 0:
            classes.append('has-hot')

        # Tooltip
        tip = _d.strftime('%A, %b %d') + ' — '
        if n == 0:
            tip += 'no leads'
        else:
            tip_parts = [str(n) + ' lead' + ('s' if n != 1 else '')]
            if hot_n: tip_parts.append(str(hot_n) + ' HOT')
            if warm_n: tip_parts.append(str(warm_n) + ' WARM')
            if cold_n: tip_parts.append(str(cold_n) + ' COLD')
            tip += ' &middot; '.join(tip_parts)

        day_events = _cell_events.get(_d, [])
        ev_n = len(day_events)
        grid_html += '<div class="' + ' '.join(classes) + '" title="' + tip + '">'
        grid_html += '<div class="cal-day-num">' + str(_d.day) + '</div>'
        if n > 0:
            dots = ''
            # Up to 6 dots total — priority HOT > WARM > COLD
            slots = 6
            n_hot = min(hot_n, slots); slots -= n_hot
            n_warm = min(warm_n, slots); slots -= n_warm
            n_cold = min(cold_n, slots)
            for _ in range(n_hot): dots += '<span class="cal-dot dot-hot"></span>'
            for _ in range(n_warm): dots += '<span class="cal-dot dot-warm"></span>'
            for _ in range(n_cold): dots += '<span class="cal-dot dot-cold"></span>'
            grid_html += '<div class="cal-dots">' + dots + '</div>'
            grid_html += '<div class="cal-count">' + str(n) + ' lead' + ('s' if n != 1 else '') + '</div>'
        elif ev_n == 0:
            grid_html += '<div class="cal-empty">—</div>'
        # Event chips — show up to 2 events per cell with overflow indicator
        if ev_n > 0:
            for _ev in day_events[:2]:
                _ev_title = (_ev.get('title') or '').upper()[:18]
                _ev_time = (_ev.get('event_time') or '')[:5]  # HH:MM
                _label = _ev_title
                if _ev_time:
                    _label = _ev_time + ' ' + _ev_title
                grid_html += '<div class="cal-event-chip">' + _label + '</div>'
            if ev_n > 2:
                grid_html += '<div class="cal-event-extra">+' + str(ev_n - 2) + ' MORE</div>'
        grid_html += '</div>'
    grid_html += '</div>'
    # Close cal-wrap
    grid_html += '</div>'
    st.markdown(grid_html, unsafe_allow_html=True)

    # ---- VIEW DAY PICKER + DAY DETAIL CARD ----
    st.markdown('<div class="view-day-wrap">', unsafe_allow_html=True)
    vd_col, vd_today_col, _vd_sp = st.columns([2, 2, 6])
    with vd_today_col:
        # Invisible faux-label to align the button vertically with the date input (which has its own label above)
        st.markdown('<div class="vd-faux-label">&nbsp;</div>', unsafe_allow_html=True)
        st.markdown('<div class="vd-today-btn-wrap">', unsafe_allow_html=True)
        if st.button('↺ TODAY', key='vd_today_btn', use_container_width=True):
            # Drop the widget's stored value so it reinitializes with our new default
            if 'view_day_picker' in st.session_state:
                del st.session_state['view_day_picker']
            st.session_state.focus_day_val = _today_d
            st.session_state.cal_year = _today_d.year
            st.session_state.cal_month = _today_d.month
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with vd_col:
        picked_day = st.date_input('// VIEW DAY', value=_focus_d, key='view_day_picker')
    st.markdown('</div>', unsafe_allow_html=True)

    # Sync focus_day state when picker changes
    if picked_day != _focus_d:
        st.session_state.focus_day_val = picked_day
        # Also switch the calendar month if picked day is in a different month
        if picked_day.year != _cyr or picked_day.month != _cmo:
            st.session_state.cal_year = picked_day.year
            st.session_state.cal_month = picked_day.month
        st.rerun()

    # Build day-detail card
    is_today = picked_day == _today_d
    day_title = picked_day.strftime('%A, %B %d')

    day_events = [_e for _e in events_all if str(_e.get('event_date', '')) == picked_day.strftime('%Y-%m-%d')]
    day_events.sort(key=lambda e: str(e.get('event_time') or '99:99'))
    day_leads = [_l for _l in leads if (_l.get('timestamp') or '').startswith(picked_day.strftime('%Y-%m-%d'))]

    card_class = 'day-detail-card' + (' is-today' if is_today else '')
    detail_html = '<div class="' + card_class + '">'
    detail_html += '<div class="day-detail-header">'
    detail_html += '<div class="day-detail-title">' + day_title + '</div>'
    if is_today:
        detail_html += '<div class="day-detail-badge">TODAY</div>'
    detail_html += '</div>'

    # Events section
    if day_events:
        detail_html += '<div class="day-detail-section">'
        detail_html += '<div class="day-detail-label">// EVENTS &middot; ' + str(len(day_events)) + '</div>'
        for _e in day_events:
            _t = (_e.get('event_time') or '')[:5]
            _title_str = _e.get('title') or ''
            _loc_str = _e.get('location') or ''
            _notes_str = _e.get('notes') or ''
            detail_html += '<div class="day-detail-event">'
            detail_html += '<div class="day-detail-event-time">' + (_t if _t else 'ALL DAY') + '</div>'
            detail_html += '<div class="day-detail-event-body">'
            detail_html += '<div class="day-detail-event-title">' + _title_str + '</div>'
            meta_bits = []
            if _loc_str:
                meta_bits.append(_loc_str.upper())
            if _e.get('lead_id'):
                _linked_l = next((l for l in leads if l.get('id') == _e['lead_id']), None)
                if _linked_l:
                    meta_bits.append('LEAD &middot; ' + (_linked_l.get('name') or '').upper())
            if meta_bits:
                detail_html += '<div class="day-detail-event-loc">' + ' &middot; '.join(meta_bits) + '</div>'
            if _notes_str:
                detail_html += '<div class="day-detail-event-notes">' + _notes_str + '</div>'
            detail_html += '</div></div>'
        detail_html += '</div>'

    # Leads received section
    if day_leads:
        detail_html += '<div class="day-detail-section">'
        detail_html += '<div class="day-detail-label">// LEADS RECEIVED &middot; ' + str(len(day_leads)) + '</div>'
        for _l in day_leads:
            _sc = _l.get('score', 'COLD')
            detail_html += '<div class="day-detail-lead">'
            detail_html += '<span class="badge score-' + _sc.lower() + '">' + _sc + '</span>'
            detail_html += '<span class="day-detail-lead-name">' + (_l.get('name') or '-') + '</span>'
            detail_html += '<span class="day-detail-lead-meta">' + (_l.get('property_type') or '') + ' &middot; $' + str(_l.get('budget', '-')) + ' &middot; ' + (_l.get('area') or '').upper() + '</span>'
            detail_html += '</div>'
        detail_html += '</div>'

    if not day_events and not day_leads:
        detail_html += '<div class="day-detail-empty">No activity on this day. Use <strong>+ ADD EVENT</strong> below to schedule something.</div>'

    detail_html += '</div>'
    st.markdown(detail_html, unsafe_allow_html=True)

# ======================================================
# AGENDA — upcoming events grouped by period
# ======================================================
if leads:
    if 'show_event_form' not in st.session_state:
        st.session_state.show_event_form = False

    # Re-fetch in case events were added (the CALENDAR block already fetched, reuse)
    try:
        _events_for_agenda = events_all
    except NameError:
        _events_for_agenda = fetch_events()

    _today_dd = datetime.now().date()
    _tomorrow_dd = _today_dd + timedelta(days=1)
    _week_end_dd = _today_dd + timedelta(days=7)

    # Group events
    groups = {'TODAY': [], 'TOMORROW': [], 'THIS WEEK': [], 'LATER': []}
    for _e in _events_for_agenda:
        try:
            _ed = datetime.strptime(str(_e.get('event_date')), '%Y-%m-%d').date()
        except:
            continue
        if _ed < _today_dd:
            continue  # skip past events
        if _ed == _today_dd:
            groups['TODAY'].append(_e)
        elif _ed == _tomorrow_dd:
            groups['TOMORROW'].append(_e)
        elif _ed <= _week_end_dd:
            groups['THIS WEEK'].append(_e)
        else:
            groups['LATER'].append(_e)

    # Sort each group by date then time
    for k in groups:
        groups[k].sort(key=lambda e: (str(e.get('event_date', '')), str(e.get('event_time') or '99:99')))

    total_upcoming = sum(len(v) for v in groups.values())
    st.markdown(sec_label('AGENDA', '<span class="section-count">' + str(total_upcoming) + ' upcoming</span>'), unsafe_allow_html=True)

    # Find the event being edited, if any
    editing_event = None
    if st.session_state.get('editing_event_id'):
        editing_event = next((e for e in _events_for_agenda if e.get('id') == st.session_state.editing_event_id), None)
        if not editing_event:
            # Stale id — clear it
            del st.session_state['editing_event_id']

    # Lead options list (used in form + EDIT handler for pre-fill)
    _lead_options = ['(no linked lead)']
    _lead_id_list = [None]
    for _l in leads:
        _lead_options.append(_l.get('name', 'Unknown') + ' · ' + str(_l.get('phone', '')))
        _lead_id_list.append(_l.get('id'))

    # Dynamic key suffix so each form session uses fresh widgets (avoids stale React state)
    if 'form_session_id' not in st.session_state:
        st.session_state.form_session_id = 0

    # We'll store the form values under non-widget keys (ev_*_val) to survive widget-key rotation
    if 'ev_date_val' not in st.session_state:
        st.session_state.ev_date_val = _today_dd
    if 'ev_time_str_val' not in st.session_state:
        st.session_state.ev_time_str_val = '10:00'
    if 'ev_title_val' not in st.session_state:
        st.session_state.ev_title_val = ''
    if 'ev_location_val' not in st.session_state:
        st.session_state.ev_location_val = ''
    if 'ev_lead_val' not in st.session_state:
        st.session_state.ev_lead_val = '(no linked lead)'
    if 'ev_notes_val' not in st.session_state:
        st.session_state.ev_notes_val = ''

    # Add/Edit Event toggle + form
    add_col, _addsp = st.columns([2, 8])
    with add_col:
        if editing_event:
            btn_label = '× CANCEL EDIT'
        elif st.session_state.show_event_form:
            btn_label = '× CANCEL'
        else:
            btn_label = '+ ADD EVENT'
        if st.button(btn_label, key='add_event_toggle'):
            if st.session_state.show_event_form:
                # Closing — reset values and edit mode, bump session id
                if 'editing_event_id' in st.session_state:
                    del st.session_state['editing_event_id']
                st.session_state.ev_date_val = _today_dd
                st.session_state.ev_time_str_val = '10:00'
                st.session_state.ev_title_val = ''
                st.session_state.ev_location_val = ''
                st.session_state.ev_lead_val = '(no linked lead)'
                st.session_state.ev_notes_val = ''
                st.session_state.form_session_id += 1
                st.session_state.show_event_form = False
            else:
                # Opening for NEW — set blank defaults, fresh widget keys
                st.session_state.ev_date_val = _today_dd
                st.session_state.ev_time_str_val = '10:00'
                st.session_state.ev_title_val = ''
                st.session_state.ev_location_val = ''
                st.session_state.ev_lead_val = '(no linked lead)'
                st.session_state.ev_notes_val = ''
                st.session_state.form_session_id += 1
                st.session_state.show_event_form = True
            st.rerun()

    if st.session_state.show_event_form:
        # Safety: ensure lead value is one of current options
        if st.session_state.ev_lead_val not in _lead_options:
            st.session_state.ev_lead_val = '(no linked lead)'

        sid = str(st.session_state.form_session_id)
        form_title = '// EDIT EVENT' if editing_event else '// NEW EVENT'
        st.markdown('<div class="add-event-panel">', unsafe_allow_html=True)
        st.markdown('<div class="add-event-title">' + form_title + '</div>', unsafe_allow_html=True)
        ef1, ef2, ef3 = st.columns([1, 1, 2])
        with ef1:
            ev_date = st.date_input('// DATE', value=st.session_state.ev_date_val, key='ev_date_' + sid)
        with ef2:
            ev_time_str = st.text_input('// TIME (HH:MM)', value=st.session_state.ev_time_str_val, placeholder='10:00', key='ev_time_' + sid)
        with ef3:
            ev_title = st.text_input('// TITLE', value=st.session_state.ev_title_val, placeholder='Visita de Sergio', key='ev_title_' + sid)

        ef4, ef5 = st.columns([1, 1])
        with ef4:
            ev_location = st.text_input('// LOCATION', value=st.session_state.ev_location_val, placeholder='Casa Urubo', key='ev_location_' + sid)
        with ef5:
            try:
                _lead_idx = _lead_options.index(st.session_state.ev_lead_val)
            except ValueError:
                _lead_idx = 0
            ev_lead_choice = st.selectbox('// LINKED LEAD (optional)', _lead_options, index=_lead_idx, key='ev_lead_' + sid)
            ev_lead_id = _lead_id_list[_lead_options.index(ev_lead_choice)]

        ev_notes = st.text_area('// NOTES', value=st.session_state.ev_notes_val, placeholder='Visita en la zona Urubo, llevar contrato y fotos del terreno...', key='ev_notes_' + sid, height=68)

        save_col, _savesp = st.columns([1, 5])
        with save_col:
            save_label = 'UPDATE EVENT' if editing_event else 'SAVE EVENT'
            if st.button(save_label, key='ev_save'):
                if not ev_title.strip():
                    st.error('Title is required.')
                else:
                    final_time = None
                    _t_str = (ev_time_str or '').strip()
                    if _t_str:
                        try:
                            _parsed_t = datetime.strptime(_t_str, '%H:%M')
                            final_time = _parsed_t.strftime('%H:%M:%S')
                        except ValueError:
                            st.warning('Time should be HH:MM (e.g., 10:30 or 14:00). Saving without time.')
                    if editing_event:
                        ok = update_event(
                            event_id=editing_event['id'],
                            event_date=ev_date.strftime('%Y-%m-%d'),
                            event_time=final_time,
                            title=ev_title.strip(),
                            location=ev_location.strip() if ev_location else None,
                            notes=ev_notes.strip() if ev_notes else None,
                            lead_id=ev_lead_id,
                        )
                    else:
                        ok = add_event(
                            event_date=ev_date.strftime('%Y-%m-%d'),
                            event_time=final_time,
                            title=ev_title.strip(),
                            location=ev_location.strip() if ev_location else None,
                            notes=ev_notes.strip() if ev_notes else None,
                            lead_id=ev_lead_id,
                        )
                    if ok:
                        st.success('Event updated.' if editing_event else 'Event saved.')
                        st.session_state.show_event_form = False
                        if 'editing_event_id' in st.session_state:
                            del st.session_state['editing_event_id']
                        # Reset stored values + bump session id so next open is fresh
                        st.session_state.ev_date_val = _today_dd
                        st.session_state.ev_time_str_val = '10:00'
                        st.session_state.ev_title_val = ''
                        st.session_state.ev_location_val = ''
                        st.session_state.ev_lead_val = '(no linked lead)'
                        st.session_state.ev_notes_val = ''
                        st.session_state.form_session_id += 1
                        st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # Render agenda groups
    if total_upcoming == 0:
        _ag_empty = '<div class="agenda-wrap">'
        _ag_empty += '<div class="empty-state-rich" style="background:transparent;border:none;padding:24px 8px 16px 8px;border-left:none;border-radius:0;">'
        _ag_empty += '<div class="empty-state-title">Nothing scheduled.</div>'
        _ag_empty += '<div class="empty-state-sub">Click <strong style="color:' + ACCENT_WARM + ';">+ ADD EVENT</strong> above to plan a visit, follow-up call, or reminder.</div>'
        _ag_empty += '</div>'
        _ag_empty += '</div>'
        st.markdown(_ag_empty, unsafe_allow_html=True)
    else:
        # Build leads-by-id map for linking
        _lead_by_id = {l.get('id'): l for l in leads}

        agenda_html = '<div class="agenda-wrap">'
        for period_name, items in groups.items():
            if not items:
                continue
            agenda_html += '<div class="agenda-group">'
            agenda_html += '<div class="agenda-period-row">'
            agenda_html += '<div class="agenda-period">// ' + period_name + '</div>'
            agenda_html += '<div class="agenda-period-count">' + str(len(items)) + ' item' + ('s' if len(items) != 1 else '') + '</div>'
            agenda_html += '</div>'
            for _e in items:
                _t = (_e.get('event_time') or '')[:5]
                _date_str = str(_e.get('event_date', ''))
                try:
                    _dobj = datetime.strptime(_date_str, '%Y-%m-%d').date()
                    _date_disp = _dobj.strftime('%b %d').upper()
                except:
                    _date_disp = _date_str

                # Time / date display
                time_html = '<div class="agenda-item-time">'
                if period_name in ('THIS WEEK', 'LATER'):
                    time_html += _date_disp
                    if _t:
                        time_html += '<br><span style="opacity:0.78;font-size:0.55rem;letter-spacing:0.14em;">' + _t + '</span>'
                    else:
                        time_html += '<br><span style="opacity:0.5;font-size:0.42rem;letter-spacing:0.14em;">ALL DAY</span>'
                else:
                    # TODAY / TOMORROW — emphasize time
                    if _t:
                        time_html += '<span style="font-size:0.7rem;">' + _t + '</span>'
                    else:
                        time_html += 'ALL DAY'
                time_html += '</div>'

                # Body
                agenda_html += '<div class="agenda-item">'
                agenda_html += time_html
                agenda_html += '<div class="agenda-item-body">'
                agenda_html += '<div class="agenda-item-title">' + (_e.get('title') or '') + '</div>'
                meta_parts = []
                if _e.get('location'):
                    meta_parts.append(_e['location'].upper())
                if _e.get('lead_id') and _lead_by_id.get(_e['lead_id']):
                    _linked = _lead_by_id[_e['lead_id']]
                    meta_parts.append('LEAD &middot; ' + _linked.get('name', '').upper())
                if meta_parts:
                    agenda_html += '<div class="agenda-item-meta">' + ' &middot; '.join(meta_parts) + '</div>'
                if _e.get('notes'):
                    agenda_html += '<div class="agenda-item-notes">' + _e['notes'] + '</div>'
                agenda_html += '</div>'  # close body
                agenda_html += '</div>'  # close item
            agenda_html += '</div>'  # close group
        agenda_html += '</div>'
        st.markdown(agenda_html, unsafe_allow_html=True)

        # Per-event edit/delete buttons (rendered as a discreet trailing row)
        with st.expander('// MANAGE EVENTS', expanded=False):
            for _e in _events_for_agenda:
                _eid = _e.get('id')
                _t = (_e.get('event_time') or '')[:5]
                _date_str = str(_e.get('event_date', ''))
                row_label = _date_str + (' ' + _t if _t else '') + '  ·  ' + (_e.get('title') or '')
                ec1, ec2, ec3 = st.columns([7, 1, 1])
                with ec1:
                    st.markdown('<div style="padding:6px 0;font-family:JetBrains Mono,monospace;font-size:0.6rem;color:' + TEXT + ';letter-spacing:0.08em;">' + row_label + '</div>', unsafe_allow_html=True)
                with ec2:
                    if st.button('EDIT', key='edit_ev_' + str(_eid)):
                        # Compute defaults from event
                        try:
                            _ed_d = datetime.strptime(str(_e.get('event_date', '')), '%Y-%m-%d').date()
                        except:
                            _ed_d = _today_dd
                        _ed_time_str = '10:00'
                        if _e.get('event_time'):
                            try:
                                _ed_time_str = str(_e['event_time'])[:5]  # 'HH:MM'
                            except:
                                pass
                        # Lead display string for the dropdown
                        _ed_lead_disp = '(no linked lead)'
                        if _e.get('lead_id'):
                            _ed_lead = next((_lc for _lc in leads if _lc.get('id') == _e['lead_id']), None)
                            if _ed_lead:
                                _ed_lead_disp = _ed_lead.get('name', 'Unknown') + ' · ' + str(_ed_lead.get('phone', ''))
                        if _ed_lead_disp not in _lead_options:
                            _ed_lead_disp = '(no linked lead)'

                        # Store under stable *_val keys, bump session id for fresh widgets
                        st.session_state.ev_date_val = _ed_d
                        st.session_state.ev_time_str_val = _ed_time_str
                        st.session_state.ev_title_val = _e.get('title') or ''
                        st.session_state.ev_location_val = _e.get('location') or ''
                        st.session_state.ev_lead_val = _ed_lead_disp
                        st.session_state.ev_notes_val = _e.get('notes') or ''
                        st.session_state.form_session_id += 1
                        st.session_state.editing_event_id = _eid
                        st.session_state.show_event_form = True
                        st.rerun()
                with ec3:
                    if st.button('DELETE', key='del_ev_' + str(_eid)):
                        delete_event(_eid)
                        # If we were editing this event, clear that too
                        if st.session_state.get('editing_event_id') == _eid:
                            del st.session_state['editing_event_id']
                            st.session_state.show_event_form = False
                        st.rerun()

# ======================================================
# DAILY STATS
# ======================================================
if leads:
    st.markdown(sec_label('TODAY'), unsafe_allow_html=True)
    d1, d2, d3 = st.columns(3)
    with d1: st.markdown(stat_card('// TODAY\'S LEADS', today_leads, 'RECEIVED TODAY'), unsafe_allow_html=True)
    with d2: st.markdown(stat_card('// PENDING CONTACT', pending_contact, 'NEED FOLLOW-UP', accent=True), unsafe_allow_html=True)
    with d3: st.markdown(stat_card('// CONVERSION RATE', conv_rate, 'HOT + WARM / TOTAL'), unsafe_allow_html=True)

# ======================================================
# TOP ZONES + FOLLOW-UPS
# ======================================================
def follow_up_panel(leads_list):
    pending = [l for l in leads_list if l.get('status', 'Nuevo') in ('Nuevo', 'Contactado')]
    pending = sorted(pending, key=lambda l: l.get('timestamp', ''), reverse=True)
    html = '<div class="chart-panel"><div class="chart-header">'
    html += '<div class="chart-title">// FOLLOW-UPS</div>'
    html += '<div class="chart-big-num" style="color:' + ACCENT_WARM + ';font-size:1.15rem;">' + str(len(pending)) + ' pending</div>'
    html += '</div>'
    if not pending:
        html += '<div style="font-family:JetBrains Mono,monospace;font-size:0.5rem;color:' + TEXT_DIM + ';letter-spacing:0.16em;padding:20px 0;text-align:center;">ALL LEADS CONTACTED.</div>'
    else:
        html += '<div class="follow-up-list">'
        for l in pending:
            st_val = l.get('status', 'Nuevo')
            sc = l.get('score', 'COLD')
            nm = l.get('name', '-')
            prop = l.get('property_type', '')
            bgt = l.get('budget', '-')
            ago = time_ago(l.get('timestamp', ''))
            html += '<div class="follow-up-row">'
            html += '<div>'
            html += '<div class="follow-up-name">' + nm + '</div>'
            html += '<div class="follow-up-meta">' + st_val.upper() + ' &middot; ' + prop.upper() + ' &middot; $' + str(bgt) + '</div>'
            html += '</div>'
            html += '<div class="follow-up-right">'
            html += '<span class="badge score-' + sc.lower() + '">' + sc + '</span>'
            if ago:
                html += '<span class="follow-up-ago">' + ago + '</span>'
            html += '</div>'
            html += '</div>'
        html += '</div>'
    html += '</div>'
    return html

if leads:
    st.markdown(sec_label('ZONES &amp; FOLLOW-UPS'), unsafe_allow_html=True)
    zones_html = zones_leaderboard(leads)
    zl, zr = st.columns(2)
    with zl:
        if zones_html:
            header = '<div class="chart-panel"><div class="chart-header">'
            header += '<div class="chart-title">// TOP ZONES</div>'
            header += '<div class="chart-big-num">by lead count</div>'
            header += '</div>'
            st.markdown(header + zones_html + '</div>', unsafe_allow_html=True)
    with zr:
        st.markdown(follow_up_panel(leads), unsafe_allow_html=True)

# ======================================================
# RESPONSE TIME
# ======================================================
timed_full = [{'name': l.get('name', '-'), 'minutes': get_minutes(l)} for l in leads if get_minutes(l) is not None]
if timed_full:
    avg = sum(t['minutes'] for t in timed_full) / len(timed_full)
    fastest = min(timed_full, key=lambda x: x['minutes'])
    slowest = max(timed_full, key=lambda x: x['minutes'])
    st.markdown(sec_label('RESPONSE TIME', '<span class="section-count">' + str(len(timed_full)) + ' contacted</span>'), unsafe_allow_html=True)
    t1, t2, t3 = st.columns(3)
    with t1: st.markdown(resp_card('// AVG RESPONSE', fmt(int(avg)), 'ACROSS ' + str(len(timed_full)) + ' LEADS'), unsafe_allow_html=True)
    with t2: st.markdown(resp_card('// FASTEST', fmt(fastest['minutes']), fastest['name'].upper(), accent=True), unsafe_allow_html=True)
    with t3: st.markdown(resp_card('// SLOWEST', fmt(slowest['minutes']), slowest['name'].upper()), unsafe_allow_html=True)

# ======================================================
# SEARCH + FILTERS
# ======================================================
st.markdown(sec_label('LEADS'), unsafe_allow_html=True)

# Build month list from leads' timestamps (most recent first)
_months_seen = set()
for _l in leads:
    _ts = (_l.get('timestamp') or '')[:7]  # YYYY-MM
    if len(_ts) == 7:
        try:
            datetime.strptime(_ts, '%Y-%m')
            _months_seen.add(_ts)
        except:
            pass
_months_sorted = sorted(_months_seen, reverse=True)
_month_label_to_key = {'All time': None}
_month_options = ['All time']
for _m in _months_sorted:
    try:
        _d = datetime.strptime(_m, '%Y-%m')
        _label = _d.strftime('%B %Y')
        _month_options.append(_label)
        _month_label_to_key[_label] = _m
    except:
        pass

search_col, f1, f2, f3 = st.columns([2, 1, 1, 1])
with search_col:
    search = st.text_input('// SEARCH', placeholder='Name or phone number', label_visibility='collapsed')
with f1:
    filter_score = st.selectbox('// SCORE', ['All', 'HOT', 'WARM', 'COLD'])
with f2:
    filter_status = st.selectbox('// STATUS', ['All', 'Nuevo', 'Contactado', 'Visitado', 'Cerrado'])
with f3:
    filter_month = st.selectbox('// MONTH', _month_options)

filtered = leads
if search.strip():
    q = search.strip().lower()
    filtered = [l for l in filtered if q in l.get('name', '').lower() or q in l.get('phone', '').lower()]
if filter_score != 'All':
    filtered = [l for l in filtered if l.get('score') == filter_score]
if filter_status != 'All':
    filtered = [l for l in filtered if l.get('status', 'Nuevo') == filter_status]
if filter_month != 'All time':
    _month_key = _month_label_to_key.get(filter_month)
    if _month_key:
        filtered = [l for l in filtered if (l.get('timestamp') or '').startswith(_month_key)]

leads_hdr_badges = ('<span class="badge score-hot" style="margin-left:6px;">HOT ' + str(hot) + '</span>'
                    '<span class="badge score-warm" style="margin-left:5px;">WARM ' + str(warm) + '</span>'
                    '<span class="badge score-cold" style="margin-left:5px;">COLD ' + str(cold) + '</span>')
st.markdown(
    '<div style="display:flex;justify-content:space-between;align-items:center;'
    'margin-bottom:10px;padding-bottom:10px;border-bottom:1px solid ' + BORDER + ';">'
    '<div style="font-family:JetBrains Mono,monospace;font-size:0.55rem;letter-spacing:0.22em;color:' + TEXT_DIM + ';">'
    '// LEADS &nbsp;&middot;&nbsp; ' + str(len(filtered)) + ' SHOWING</div>'
    '<div>' + leads_hdr_badges + '</div>'
    '</div>',
    unsafe_allow_html=True
)

# ======================================================
# LEAD CARDS
# ======================================================
if not filtered:
    _has_filters_active = bool(search.strip()) or filter_score != 'All' or filter_status != 'All' or filter_month != 'All time'
    if total == 0:
        # Truly empty — no leads ever
        _empty_html = '<div class="empty-state-rich">'
        _empty_html += '<div class="empty-state-title">No leads yet.</div>'
        _empty_html += '<div class="empty-state-sub">New leads will appear here automatically as they come in through the chat assistant.</div>'
        _empty_html += '</div>'
        st.markdown(_empty_html, unsafe_allow_html=True)
    elif _has_filters_active:
        _empty_html = '<div class="empty-state-rich">'
        _empty_html += '<div class="empty-state-title">Nothing matches.</div>'
        _empty_html += '<div class="empty-state-sub">Adjust the filters above to see more leads, or clear them to see everything.</div>'
        _empty_html += '</div>'
        st.markdown(_empty_html, unsafe_allow_html=True)
    else:
        _empty_html = '<div class="empty-state-rich">'
        _empty_html += '<div class="empty-state-title">No leads to show.</div>'
        _empty_html += '<div class="empty-state-sub">Once you have leads they will appear here.</div>'
        _empty_html += '</div>'
        st.markdown(_empty_html, unsafe_allow_html=True)
else:
    statuses_list = ['Nuevo', 'Contactado', 'Visitado', 'Cerrado']
    score_md_map = {'HOT': ':red[**HOT**]', 'WARM': ':orange[**WARM**]', 'COLD': ':gray[**COLD**]'}
    for lead in filtered:
        lead_id = lead.get('id')
        score = lead.get('score', 'COLD')
        status = lead.get('status', 'Nuevo')
        name = lead.get('name', '-')
        phone = lead.get('phone', '').replace('+', '').replace(' ', '').replace('-', '')
        ptype = lead.get('property_type', '')
        area = lead.get('area', '')
        budget = lead.get('budget', '-')
        timeline = lead.get('timeline', '-')
        ts = lead.get('timestamp', '-')
        minutes = get_minutes(lead)
        ago = time_ago(lead.get('timestamp', ''))

        # Markdown label for the expander summary
        score_md = score_md_map.get(score, ':gray[COLD]')
        # Compact parts: name (bold) · SCORE (colored) · STATUS · $budget · zone · ago
        label_parts = ['**' + name + '**', score_md, status.upper(), '$' + str(budget)]
        if area:
            label_parts.append(area.upper())
        if ago:
            label_parts.append(ago)
        lead_label = '   ·   '.join(label_parts)

        wa_msg = 'Hola ' + name + ', soy de la agencia inmobiliaria LeadBoost. Te contactamos porque mostraste interes en ' + ptype + ' en ' + area + '. Tienes un momento para hablar?'
        wa_link = 'https://wa.me/591' + phone + '?text=' + quote(wa_msg)

        # Apply a per-lead class via a wrapper div around the expander
        st.markdown('<div class="lead-expander-wrap lead-' + score.lower() + '">', unsafe_allow_html=True)
        with st.expander(lead_label, expanded=False):
            # Inner header: response time pill if available
            resp_html = ''
            if minutes is not None:
                resp_html = '<span class="resp-time" style="margin-right:8px;">' + fmt(minutes) + ' RESPONSE</span>'
            if resp_html:
                st.markdown('<div style="margin-bottom:6px;">' + resp_html + '</div>', unsafe_allow_html=True)

            # Field grid
            grid_html = '<div class="lead-grid">'
            grid_html += '<div><span class="field-label">// PHONE</span>' + str(lead.get('phone', '-')) + '</div>'
            grid_html += '<div><span class="field-label">// PROPERTY</span>' + ptype + '</div>'
            grid_html += '<div><span class="field-label">// LOCATION</span>' + area + '</div>'
            grid_html += '<div><span class="field-label">// BUDGET</span>$' + str(budget) + '</div>'
            grid_html += '<div><span class="field-label">// TIMELINE</span>' + str(timeline) + ' meses</div>'
            grid_html += '<div><span class="field-label">// SCORE</span>' + score + '</div>'
            grid_html += '</div>'
            st.markdown(grid_html, unsafe_allow_html=True)

            # Footer with received timestamp + WhatsApp button
            footer_html = '<div class="lead-footer" style="margin-top:14px;padding-top:12px;border-top:1px solid ' + BORDER + ';">'
            footer_html += '<span class="lead-ts">// RECEIVED &middot; ' + str(ts) + '</span>'
            footer_html += '<a class="wa-btn" href="' + wa_link + '" target="_blank">// WHATSAPP</a>'
            footer_html += '</div>'
            st.markdown(footer_html, unsafe_allow_html=True)

            # Status buttons
            st.markdown('<div style="font-family:JetBrains Mono,monospace;font-size:0.52rem;letter-spacing:0.22em;color:' + TEXT_DIM + ';text-transform:uppercase;margin-top:16px;margin-bottom:6px;">// STATUS</div>', unsafe_allow_html=True)
            status_cols = st.columns(4)
            for si, s in enumerate(statuses_list):
                with status_cols[si]:
                    is_current = status == s
                    if st.button(s.upper(), key='s_' + str(lead_id) + '_' + s, disabled=is_current):
                        update_status(lead_id, s)
                        st.rerun()

            # Notes
            current_note = lead.get('notes') or ''
            note_input = st.text_area(
                '// NOTES',
                value=current_note,
                key='note_' + str(lead_id),
                placeholder='Add agent notes here...',
                height=64,
                label_visibility='visible'
            )
            if st.button('SAVE NOTE', key='save_note_' + str(lead_id)):
                try:
                    get_supabase().table('leads').update({'notes': note_input}).eq('id', lead_id).execute()
                    st.success('Saved.')
                    st.rerun()
                except Exception as e:
                    st.error('Error: ' + str(e))
        st.markdown('</div>', unsafe_allow_html=True)

# ======================================================
# LOGOUT + FOOTER
# ======================================================
st.markdown('<div style="border-top:1px solid ' + BORDER + ';margin-top:28px;padding-top:18px;"></div>', unsafe_allow_html=True)
foot_l, foot_r = st.columns([2, 5])
with foot_l:
    if st.button('// LOG OUT'):
        st.session_state.authenticated = False
        st.rerun()
with foot_r:
    today_ver = datetime.now().strftime('%Y-%m-%d')
    st.markdown(
        '<div style="padding-top:10px;font-family:JetBrains Mono,monospace;font-size:0.48rem;'
        'color:' + TEXT_DIM + ';letter-spacing:0.14em;text-align:right;">'
        'LEADBOOST OS &middot; V3.0 &middot; ' + today_ver + '</div>',
        unsafe_allow_html=True
    )
